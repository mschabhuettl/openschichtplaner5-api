"""Export Scheduler router for OpenSchichtplaner5.

Provides CRUD endpoints for weekly export schedule configurations,
plus a manual trigger endpoint that generates and emails a report.

Endpoints:
  GET    /api/export-scheduler/schedules          – list schedules (Planer+)
  POST   /api/export-scheduler/schedules          – create schedule (Admin)
  PUT    /api/export-scheduler/schedules/{id}     – update schedule (Admin)
  DELETE /api/export-scheduler/schedules/{id}     – delete schedule (Admin)
  POST   /api/export-scheduler/schedules/{id}/run – trigger manually (Admin)
"""

from __future__ import annotations

import calendar as _calendar
import io
import json
import re
import uuid
from datetime import UTC
from datetime import datetime as _dt
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field, field_validator

from ..dependencies import require_admin, require_planer

router = APIRouter(tags=["Export Scheduler"])

# ── Storage ────────────────────────────────────────────────────────────────────

_DATA_DIR = Path(__file__).parent.parent.parent / "data"
_DATA_DIR.mkdir(parents=True, exist_ok=True)
_SCHEDULES_FILE = _DATA_DIR / "export_schedules.json"


def _load_schedules() -> list[dict]:
    """Load schedules from JSON file."""
    if not _SCHEDULES_FILE.exists():
        return []
    try:
        with open(_SCHEDULES_FILE, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except (json.JSONDecodeError, OSError):
        return []


def _save_schedules(schedules: list[dict]) -> None:
    """Persist schedules to JSON file."""
    with open(_SCHEDULES_FILE, "w", encoding="utf-8") as f:
        json.dump(schedules, f, ensure_ascii=False, indent=2)


def _get_schedule_by_id(schedule_id: str) -> tuple[list[dict], int, dict]:
    """Return (schedules, index, item) or raise 404."""
    schedules = _load_schedules()
    for i, s in enumerate(schedules):
        if s.get("id") == schedule_id:
            return schedules, i, s
    raise HTTPException(status_code=404, detail=f"Schedule '{schedule_id}' not found")


# ── Models ─────────────────────────────────────────────────────────────────────

_TIME_RE = re.compile(r"^\d{2}:\d{2}$")


class ScheduleCreate(BaseModel):
    """Request body for creating a new export schedule."""

    name: str = Field(..., min_length=1, max_length=100, description="Human-readable schedule name")
    frequency: str = Field("weekly", description="Only 'weekly' supported")
    day_of_week: int = Field(..., ge=0, le=6, description="0=Monday … 6=Sunday")
    time: str = Field(..., description="Time in HH:MM format")
    format: str = Field(..., description="'xlsx' or 'csv'")
    group_id: int | None = Field(None, description="Optional group filter")
    email_to: list[str] = Field(..., min_length=1, description="List of recipient email addresses")
    enabled: bool = Field(True, description="Whether the schedule is active")

    @field_validator("frequency")
    @classmethod
    def validate_frequency(cls, v: str) -> str:
        if v != "weekly":
            raise ValueError("Only 'weekly' frequency is supported")
        return v

    @field_validator("format")
    @classmethod
    def validate_format(cls, v: str) -> str:
        if v not in ("xlsx", "csv"):
            raise ValueError("format must be 'xlsx' or 'csv'")
        return v

    @field_validator("time")
    @classmethod
    def validate_time(cls, v: str) -> str:
        if not _TIME_RE.match(v):
            raise ValueError("time must be in HH:MM format")
        h, m = int(v[:2]), int(v[3:])
        if h > 23 or m > 59:
            raise ValueError("time must be a valid HH:MM (00:00–23:59)")
        return v

    @field_validator("email_to")
    @classmethod
    def validate_emails(cls, v: list[str]) -> list[str]:
        if not v:
            raise ValueError("email_to must contain at least one address")
        for addr in v:
            if "@" not in addr:
                raise ValueError(f"Invalid email address: {addr}")
        return v


class ScheduleUpdate(BaseModel):
    """Request body for updating an existing export schedule (all fields optional)."""

    name: str | None = Field(None, min_length=1, max_length=100)
    frequency: str | None = None
    day_of_week: int | None = Field(None, ge=0, le=6)
    time: str | None = None
    format: str | None = None
    group_id: int | None = None
    email_to: list[str] | None = None
    enabled: bool | None = None

    @field_validator("frequency")
    @classmethod
    def validate_frequency(cls, v: str | None) -> str | None:
        if v is not None and v != "weekly":
            raise ValueError("Only 'weekly' frequency is supported")
        return v

    @field_validator("format")
    @classmethod
    def validate_format(cls, v: str | None) -> str | None:
        if v is not None and v not in ("xlsx", "csv"):
            raise ValueError("format must be 'xlsx' or 'csv'")
        return v

    @field_validator("time")
    @classmethod
    def validate_time(cls, v: str | None) -> str | None:
        if v is not None:
            if not _TIME_RE.match(v):
                raise ValueError("time must be in HH:MM format")
            h, m = int(v[:2]), int(v[3:])
            if h > 23 or m > 59:
                raise ValueError("time must be a valid HH:MM (00:00–23:59)")
        return v

    @field_validator("email_to")
    @classmethod
    def validate_emails(cls, v: list[str] | None) -> list[str] | None:
        if v is not None:
            if not v:
                raise ValueError("email_to must contain at least one address")
            for addr in v:
                if "@" not in addr:
                    raise ValueError(f"Invalid email address: {addr}")
        return v


# ── Helpers ────────────────────────────────────────────────────────────────────


def _generate_export(fmt: str, group_id: int | None, month: str) -> tuple[bytes, int]:
    """Generate an Excel or CSV export for the given month.

    Returns (file_bytes, row_count).
    """
    try:
        dt = _dt.strptime(month, "%Y-%m")
        year, mon = dt.year, dt.month
    except ValueError as exc:
        raise ValueError(f"Invalid month format: {month}") from exc

    from sp5lib.db import get_db

    db = get_db()
    entries = db.get_schedule(year=year, month=mon, group_id=group_id)
    employees = db.get_employees(include_hidden=False)

    if group_id is not None:
        member_ids = set(db.get_group_members(group_id))
        employees = [e for e in employees if e["ID"] in member_ids]
    employees.sort(key=lambda x: x.get("POSITION", 0))

    entry_map: dict = {}
    for entry in entries:
        key = (entry["employee_id"], entry["date"])
        entry_map[key] = entry

    num_days = _calendar.monthrange(year, mon)[1]
    days = [f"{year:04d}-{mon:02d}-{d:02d}" for d in range(1, num_days + 1)]
    row_count = len(employees)

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError("openpyxl is not installed") from exc

        _month_names_de = [
            "Januar", "Februar", "März", "April", "Mai", "Juni",
            "Juli", "August", "September", "Oktober", "November", "Dezember",
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{_month_names_de[mon - 1]} {year}"
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(bold=True, color="FFFFFF", size=9)
        header_fill = PatternFill(fill_type="solid", fgColor="1E293B")

        ws.cell(1, 1, "Mitarbeiter").font = header_font
        ws.cell(1, 1).fill = header_fill
        ws.cell(1, 1).alignment = Alignment(horizontal="left")
        ws.cell(1, 1).border = border
        ws.column_dimensions["A"].width = 22
        ws.cell(1, 2, "Kürzel").font = header_font
        ws.cell(1, 2).fill = header_fill
        ws.cell(1, 2).alignment = Alignment(horizontal="center")
        ws.cell(1, 2).border = border
        ws.column_dimensions["B"].width = 6

        for d in range(1, num_days + 1):
            col = d + 2
            wd = _dt(year, mon, d).weekday()
            wd_abbr = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"][wd]
            cell = ws.cell(1, col, f"{d}\n{wd_abbr}")
            cell.font = header_font
            is_weekend = wd >= 5
            cell.fill = PatternFill(fill_type="solid", fgColor="475569" if is_weekend else "1E293B")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 4.5
        ws.row_dimensions[1].height = 28

        for r_idx, emp in enumerate(employees, start=2):
            emp_name = f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(", ")
            short = emp.get("SHORTNAME", "")
            cbklabel = emp.get("CBKLABEL", 16777215)
            cbklabel_hex = emp.get("CBKLABEL_HEX", "#f8fafc")
            cfglabel_hex = emp.get("CFGLABEL_HEX", "#000000")
            emp_bg = (
                cbklabel_hex.lstrip("#")
                if (cbklabel and cbklabel != 16777215 and cbklabel != 0)
                else "F8FAFC"
            )
            emp_fg = (
                cfglabel_hex.lstrip("#")
                if (cbklabel and cbklabel != 16777215 and cbklabel != 0)
                else "1E293B"
            )
            name_cell = ws.cell(r_idx, 1, emp_name)
            name_cell.font = Font(bold=bool(emp.get("BOLD")), color=emp_fg, size=9)
            name_cell.fill = PatternFill(fill_type="solid", fgColor=emp_bg)
            name_cell.border = border
            short_cell = ws.cell(r_idx, 2, short)
            short_cell.font = Font(color="64748B", size=9)
            short_cell.fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
            short_cell.alignment = Alignment(horizontal="center")
            short_cell.border = border
            for d in range(1, num_days + 1):
                col = d + 2
                date_str = f"{year:04d}-{mon:02d}-{d:02d}"
                wd = _dt(year, mon, d).weekday()
                is_weekend = wd >= 5
                e = entry_map.get((emp["ID"], date_str))
                cell = ws.cell(r_idx, col)
                if e:
                    bg = e.get("color_bk", "#4A90D9").lstrip("#")
                    fg = e.get("color_text", "#FFFFFF").lstrip("#")
                    cell.value = e.get("display_name", "")
                    cell.font = Font(bold=True, color=fg, size=8)
                    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                else:
                    cell.fill = PatternFill(fill_type="solid", fgColor="EBEBEB" if is_weekend else "FFFFFF")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            ws.row_dimensions[r_idx].height = 14
        ws.freeze_panes = "C2"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), row_count

    # CSV
    import csv as _csv

    rows = []
    for emp in employees:
        row: dict[str, Any] = {
            "Mitarbeiter": f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(", "),
            "Kürzel": emp.get("SHORTNAME", ""),
        }
        for date in days:
            day_num = int(date.split("-")[2])
            e = entry_map.get((emp["ID"], date))
            row[str(day_num)] = e["display_name"] if e else ""
        rows.append(row)

    buf2 = io.StringIO()
    if rows:
        writer = _csv.DictWriter(buf2, fieldnames=rows[0].keys(), lineterminator="\r\n")
        writer.writeheader()
        writer.writerows(rows)
    return buf2.getvalue().encode("utf-8-sig"), row_count


def _send_export_email(
    schedule: dict,
    file_bytes: bytes,
    row_count: int,
    month: str,
    fmt: str,
) -> dict:
    """Send the export file via email.  Returns result dict."""
    import smtplib
    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    from sp5lib.email_service import get_config

    cfg = get_config()
    if not cfg.is_configured:
        return {
            "success": False,
            "reason": "SMTP not configured",
            "export_url": f"/api/export/schedule?month={month}&format={fmt}",
        }

    ext = "xlsx" if fmt == "xlsx" else "csv"
    filename = f"dienstplan_{month}.{ext}"
    subject = f"[SP5] Dienstplan-Export {month} – {schedule.get('name', '')}"
    body_plain = (
        f"Automatisch generierter Dienstplan-Export für {month}.\n\n"
        f"Schedule: {schedule.get('name', '')}\n"
        f"Zeilen: {row_count}\n"
        f"Format: {fmt.upper()}\n\n"
        "---\nOpenSchichtplaner5"
    )
    body_html = f"""\
<!DOCTYPE html><html><head><meta charset="utf-8"></head><body
  style="font-family:sans-serif;color:#1e293b;">
<h2>📋 Dienstplan-Export {month}</h2>
<p>Schedule: <strong>{schedule.get('name', '')}</strong></p>
<p>Zeilen: {row_count} &nbsp;|&nbsp; Format: {fmt.upper()}</p>
<p style="color:#64748b;font-size:12px;">
  Automatisch generiert von OpenSchichtplaner5
</p>
</body></html>"""

    mime_type = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if fmt == "xlsx"
        else "text/csv"
    )

    sent_to: list[str] = []
    failed: list[str] = []

    for recipient in schedule.get("email_to", []):
        msg = MIMEMultipart("mixed")
        msg["From"] = cfg.from_addr
        msg["To"] = recipient
        msg["Subject"] = subject

        alt = MIMEMultipart("alternative")
        alt.attach(MIMEText(body_plain, "plain", "utf-8"))
        alt.attach(MIMEText(body_html, "html", "utf-8"))
        msg.attach(alt)

        attachment = MIMEBase(*mime_type.split("/", 1))
        attachment.set_payload(file_bytes)
        encoders.encode_base64(attachment)
        attachment.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(attachment)

        try:
            if cfg.tls_mode == "ssl":
                with smtplib.SMTP_SSL(cfg.host, cfg.port, timeout=15) as srv:
                    if cfg.user:
                        srv.login(cfg.user, cfg.password)
                    srv.send_message(msg)
            else:
                with smtplib.SMTP(cfg.host, cfg.port, timeout=15) as srv:
                    if cfg.tls_mode in ("true", "starttls", "1", "yes"):
                        srv.starttls()
                    if cfg.user:
                        srv.login(cfg.user, cfg.password)
                    srv.send_message(msg)
            sent_to.append(recipient)
        except Exception:
            failed.append(recipient)

    if sent_to:
        return {
            "success": True,
            "sent_to": sent_to,
            "failed": failed,
            "format": fmt,
            "rows": row_count,
        }
    return {
        "success": False,
        "reason": "All email sends failed",
        "failed": failed,
        "format": fmt,
        "rows": row_count,
    }


# ── Endpoints ──────────────────────────────────────────────────────────────────


@router.get(
    "/api/export-scheduler/schedules",
    summary="List export schedules",
    description="Returns all configured export schedules. **Required role:** Planer",
)
def list_schedules(_cur_user: dict = Depends(require_planer)) -> list[dict]:
    """Return all export schedules."""
    return _load_schedules()


@router.post(
    "/api/export-scheduler/schedules",
    summary="Create export schedule",
    description="Create a new weekly export schedule. **Required role:** Admin",
    status_code=201,
)
def create_schedule(
    body: ScheduleCreate,
    _cur_user: dict = Depends(require_admin),
) -> dict:
    """Create a new export schedule."""
    schedules = _load_schedules()
    new_schedule = {
        "id": str(uuid.uuid4()),
        "name": body.name,
        "frequency": body.frequency,
        "day_of_week": body.day_of_week,
        "time": body.time,
        "format": body.format,
        "group_id": body.group_id,
        "email_to": body.email_to,
        "enabled": body.enabled,
        "created_at": _dt.now(UTC).isoformat() + "Z",
        "updated_at": _dt.now(UTC).isoformat() + "Z",
    }
    schedules.append(new_schedule)
    _save_schedules(schedules)
    return new_schedule


@router.put(
    "/api/export-scheduler/schedules/{schedule_id}",
    summary="Update export schedule",
    description="Update an existing export schedule. **Required role:** Admin",
)
def update_schedule(
    schedule_id: str,
    body: ScheduleUpdate,
    _cur_user: dict = Depends(require_admin),
) -> dict:
    """Update an existing export schedule."""
    schedules, idx, existing = _get_schedule_by_id(schedule_id)
    update_data = body.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        existing[key] = value
    existing["updated_at"] = _dt.now(UTC).isoformat() + "Z"
    schedules[idx] = existing
    _save_schedules(schedules)
    return existing


@router.delete(
    "/api/export-scheduler/schedules/{schedule_id}",
    summary="Delete export schedule",
    description="Delete an export schedule permanently. **Required role:** Admin",
    status_code=204,
)
def delete_schedule(
    schedule_id: str,
    _cur_user: dict = Depends(require_admin),
) -> None:
    """Delete an export schedule."""
    schedules, idx, _ = _get_schedule_by_id(schedule_id)
    schedules.pop(idx)
    _save_schedules(schedules)


@router.post(
    "/api/export-scheduler/schedules/{schedule_id}/run",
    summary="Manually trigger an export schedule",
    description=(
        "Generate the export report immediately and send it via email.\n\n"
        "If SMTP is not configured, returns an export URL instead.\n\n"
        "**Required role:** Admin"
    ),
)
def run_schedule(
    schedule_id: str,
    _cur_user: dict = Depends(require_admin),
) -> dict:
    """Manually trigger a schedule: generate export and send email."""
    _, _, schedule = _get_schedule_by_id(schedule_id)

    month = _dt.now(UTC).strftime("%Y-%m")
    fmt = schedule.get("format", "xlsx")
    group_id = schedule.get("group_id")

    try:
        file_bytes, row_count = _generate_export(fmt, group_id, month)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Export generation failed: {exc}") from exc

    return _send_export_email(schedule, file_bytes, row_count, month, fmt)
