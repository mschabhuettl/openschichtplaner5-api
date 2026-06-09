"""Scheduled Reports router for OpenSchichtplaner5.

Provides CRUD endpoints for scheduled report configurations plus a background
scheduler that automatically generates and emails reports on configured frequencies.

Supported report types:
  - schedule_overview   : Monthly schedule grid (Excel/CSV)
  - overtime            : Overtime/underhours summary
  - absences            : Absence statistics

Endpoints:
  GET    /api/scheduled-reports              – list reports (Planer+)
  POST   /api/scheduled-reports              – create report (Admin)
  GET    /api/scheduled-reports/{id}         – get single report (Planer+)
  PUT    /api/scheduled-reports/{id}         – update report (Admin)
  DELETE /api/scheduled-reports/{id}         – delete report (Admin)
  POST   /api/scheduled-reports/{id}/run     – trigger manually (Admin)
  GET    /api/scheduled-reports/scheduler/status – scheduler status (Admin)

  All endpoints also available under /api/v1/scheduled-reports/ (versioned alias).
"""

from __future__ import annotations

import calendar as _calendar
import csv as _csv
import io
import json
import logging
import threading
import time
import uuid
from datetime import UTC, timedelta
from datetime import datetime as _dt
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Depends, HTTPException
from pydantic import BaseModel, Field, field_validator

from ..dependencies import require_admin, require_planer

_logger = logging.getLogger("sp5.scheduled_reports")

router = APIRouter(prefix="/api/scheduled-reports", tags=["Scheduled Reports"])

# ── Storage ────────────────────────────────────────────────────────────────────

_DATA_DIR = Path(__file__).parent.parent.parent / "data"
_DATA_DIR.mkdir(parents=True, exist_ok=True)
_REPORTS_FILE = _DATA_DIR / "scheduled_reports.json"

VALID_REPORT_TYPES = {"schedule_overview", "overtime", "absences"}
VALID_FREQUENCIES = {"daily", "weekly", "monthly"}
VALID_FORMATS = {"xlsx", "csv"}


def _load_reports() -> list[dict]:
    """Load scheduled reports from JSON file."""
    if not _REPORTS_FILE.exists():
        return []
    try:
        with open(_REPORTS_FILE, encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except (json.JSONDecodeError, OSError):
        return []


def _save_reports(reports: list[dict]) -> None:
    """Persist scheduled reports to JSON file."""
    with open(_REPORTS_FILE, "w", encoding="utf-8") as f:
        json.dump(reports, f, ensure_ascii=False, indent=2, default=str)


def _get_report_by_id(report_id: str) -> tuple[list[dict], int, dict]:
    """Return (reports, index, item) or raise 404."""
    reports = _load_reports()
    for i, r in enumerate(reports):
        if r.get("id") == report_id:
            return reports, i, r
    raise HTTPException(status_code=404, detail=f"Scheduled report '{report_id}' not found")


# ── Models ─────────────────────────────────────────────────────────────────────


class ScheduledReportCreate(BaseModel):
    """Request body for creating a scheduled report."""

    name: str = Field(..., min_length=1, max_length=100, description="Human-readable report name")
    report_type: str = Field(..., description="Type of report: schedule_overview, overtime, absences")
    frequency: str = Field(..., description="How often to send: daily, weekly, monthly")
    recipients: list[str] = Field(..., min_length=1, description="List of recipient email addresses")
    format: str = Field("xlsx", description="Output format: xlsx or csv")
    filters: dict[str, Any] = Field(default_factory=dict, description="Optional filters (group_id, etc.)")
    enabled: bool = Field(True, description="Whether the schedule is active")

    @field_validator("report_type")
    @classmethod
    def validate_report_type(cls, v: str) -> str:
        if v not in VALID_REPORT_TYPES:
            raise ValueError(f"report_type must be one of {sorted(VALID_REPORT_TYPES)}")
        return v

    @field_validator("frequency")
    @classmethod
    def validate_frequency(cls, v: str) -> str:
        if v not in VALID_FREQUENCIES:
            raise ValueError(f"frequency must be one of {sorted(VALID_FREQUENCIES)}")
        return v

    @field_validator("format")
    @classmethod
    def validate_format(cls, v: str) -> str:
        if v not in VALID_FORMATS:
            raise ValueError(f"format must be one of {sorted(VALID_FORMATS)}")
        return v

    @field_validator("recipients")
    @classmethod
    def validate_recipients(cls, v: list[str]) -> list[str]:
        if not v:
            raise ValueError("recipients must contain at least one address")
        for addr in v:
            if "@" not in addr:
                raise ValueError(f"Invalid email address: {addr}")
        return v


class ScheduledReportUpdate(BaseModel):
    """Request body for updating a scheduled report (all fields optional)."""

    name: str | None = Field(None, min_length=1, max_length=100)
    report_type: str | None = None
    frequency: str | None = None
    recipients: list[str] | None = None
    format: str | None = None
    filters: dict[str, Any] | None = None
    enabled: bool | None = None

    @field_validator("report_type")
    @classmethod
    def validate_report_type(cls, v: str | None) -> str | None:
        if v is not None and v not in VALID_REPORT_TYPES:
            raise ValueError(f"report_type must be one of {sorted(VALID_REPORT_TYPES)}")
        return v

    @field_validator("frequency")
    @classmethod
    def validate_frequency(cls, v: str | None) -> str | None:
        if v is not None and v not in VALID_FREQUENCIES:
            raise ValueError(f"frequency must be one of {sorted(VALID_FREQUENCIES)}")
        return v

    @field_validator("format")
    @classmethod
    def validate_format(cls, v: str | None) -> str | None:
        if v is not None and v not in VALID_FORMATS:
            raise ValueError(f"format must be one of {sorted(VALID_FORMATS)}")
        return v

    @field_validator("recipients")
    @classmethod
    def validate_recipients(cls, v: list[str] | None) -> list[str] | None:
        if v is not None:
            if not v:
                raise ValueError("recipients must contain at least one address")
            for addr in v:
                if "@" not in addr:
                    raise ValueError(f"Invalid email address: {addr}")
        return v


# ── Report generation helpers ──────────────────────────────────────────────────


def _compute_next_run(frequency: str, from_dt: _dt | None = None) -> str:
    """Compute the next run time for a given frequency."""
    now = from_dt or _dt.now(UTC)
    if frequency == "daily":
        next_run = now + timedelta(days=1)
    elif frequency == "weekly":
        next_run = now + timedelta(weeks=1)
    else:  # monthly
        # same day next month (clamped to month end)
        year = now.year + (now.month // 12)
        month = (now.month % 12) + 1
        max_day = _calendar.monthrange(year, month)[1]
        next_run = now.replace(year=year, month=month, day=min(now.day, max_day))
    return next_run.isoformat()


def _get_reference_month(frequency: str) -> tuple[int, int]:
    """Return (year, month) for the report period based on frequency."""
    now = _dt.now(UTC)
    if frequency == "daily":
        # yesterday
        yesterday = now - timedelta(days=1)
        return yesterday.year, yesterday.month
    else:
        # previous month
        first = now.replace(day=1)
        prev = first - timedelta(days=1)
        return prev.year, prev.month


def _generate_schedule_overview(year: int, month: int, filters: dict, fmt: str) -> tuple[bytes, str]:
    """Generate a schedule overview report as xlsx or csv."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        _HAS_OPENPYXL = True
    except ImportError:
        _HAS_OPENPYXL = False

    from sp5lib.db import get_db
    db = get_db()

    group_id = filters.get("group_id")
    entries = db.get_schedule(year=year, month=month, group_id=group_id)
    employees = db.get_employees(include_hidden=False)
    if group_id is not None:
        member_ids = set(db.get_group_members(group_id))
        employees = [e for e in employees if e["ID"] in member_ids]
    employees.sort(key=lambda x: x.get("POSITION", 0))

    entry_map: dict = {}
    for entry in entries:
        key = (entry["employee_id"], entry["date"])
        entry_map[key] = entry

    num_days = _calendar.monthrange(year, month)[1]
    days = [f"{year:04d}-{month:02d}-{d:02d}" for d in range(1, num_days + 1)]

    month_name = _dt(year, month, 1).strftime("%B %Y")
    filename = f"dienstplan_{year:04d}-{month:02d}.{fmt}"

    if fmt == "xlsx" and _HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{month_name}"
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hf = Font(bold=True, color="FFFFFF", size=9)
        hfill = PatternFill(fill_type="solid", fgColor="1E293B")

        ws.cell(1, 1, "Mitarbeiter").font = hf
        ws.cell(1, 1).fill = hfill
        ws.cell(1, 1).alignment = Alignment(horizontal="left")
        ws.cell(1, 1).border = border
        ws.column_dimensions["A"].width = 22
        ws.cell(1, 2, "Kürzel").font = hf
        ws.cell(1, 2).fill = hfill
        ws.cell(1, 2).alignment = Alignment(horizontal="center")
        ws.cell(1, 2).border = border
        ws.column_dimensions["B"].width = 6

        for d in range(1, num_days + 1):
            col = d + 2
            wd = _dt(year, month, d).weekday()
            wd_abbr = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"][wd]
            cell = ws.cell(1, col, f"{d}\n{wd_abbr}")
            cell.font = hf
            is_weekend = wd >= 5
            cell.fill = PatternFill(fill_type="solid", fgColor="475569" if is_weekend else "1E293B")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 4.5
        ws.row_dimensions[1].height = 28

        for r_idx, emp in enumerate(employees, start=2):
            emp_name = f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(", ")
            short = emp.get("SHORTNAME", "")
            cbklabel = emp.get("CBKLABEL", 16777215)
            cbklabel_hex = emp.get("CBKLABEL_HEX", "#f8fafc")
            cfglabel_hex = emp.get("CFGLABEL_HEX", "#000000")
            emp_bg = (
                cbklabel_hex.lstrip("#")
                if (cbklabel and cbklabel not in (0, 16777215))
                else "F8FAFC"
            )
            emp_fg = (
                cfglabel_hex.lstrip("#")
                if (cbklabel and cbklabel not in (0, 16777215))
                else "1E293B"
            )
            name_cell = ws.cell(r_idx, 1, emp_name)
            name_cell.font = Font(bold=bool(emp.get("BOLD")), color=emp_fg, size=9)
            name_cell.fill = PatternFill(fill_type="solid", fgColor=emp_bg)
            name_cell.border = border
            short_cell = ws.cell(r_idx, 2, short)
            short_cell.font = Font(color="64748B", size=9)
            short_cell.fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
            short_cell.alignment = Alignment(horizontal="center")
            short_cell.border = border
            for d in range(1, num_days + 1):
                col = d + 2
                date_str = f"{year:04d}-{month:02d}-{d:02d}"
                wd = _dt(year, month, d).weekday()
                is_weekend = wd >= 5
                e = entry_map.get((emp["ID"], date_str))
                cell = ws.cell(r_idx, col)
                if e:
                    bg = e.get("color_bk", "#4A90D9").lstrip("#")
                    fg = e.get("color_text", "#FFFFFF").lstrip("#")
                    cell.value = e.get("display_name", "")
                    cell.font = Font(bold=True, color=fg, size=8)
                    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                else:
                    cell.fill = PatternFill(fill_type="solid", fgColor="EBEBEB" if is_weekend else "FFFFFF")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            ws.row_dimensions[r_idx].height = 14
        ws.freeze_panes = "C2"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), filename

    # CSV fallback
    rows = []
    for emp in employees:
        row: dict[str, Any] = {
            "Mitarbeiter": f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(", "),
            "Kürzel": emp.get("SHORTNAME", ""),
        }
        for date_str in days:
            day_num = int(date_str.split("-")[2])
            e = entry_map.get((emp["ID"], date_str))
            row[str(day_num)] = e["display_name"] if e else ""
        rows.append(row)

    buf2 = io.StringIO()
    if rows:
        writer = _csv.DictWriter(buf2, fieldnames=list(rows[0].keys()), lineterminator="\r\n")
        writer.writeheader()
        writer.writerows(rows)
    csv_filename = f"dienstplan_{year:04d}-{month:02d}.csv"
    return buf2.getvalue().encode("utf-8-sig"), csv_filename


def _generate_overtime_report(year: int, month: int, filters: dict, fmt: str) -> tuple[bytes, str]:
    """Generate an overtime/underhours summary report."""
    from sp5lib.db import get_db
    db = get_db()

    group_id = filters.get("group_id")
    employees = db.get_employees(include_hidden=False)
    if group_id is not None:
        member_ids = set(db.get_group_members(group_id))
        employees = [e for e in employees if e["ID"] in member_ids]
    employees.sort(key=lambda x: x.get("POSITION", 0))

    num_days = _calendar.monthrange(year, month)[1]
    working_days = sum(
        1 for d in range(1, num_days + 1)
        if _dt(year, month, d).weekday() < 5
    )

    rows = []
    for emp in employees:
        emp_id = emp["ID"]
        entries = db.get_schedule(year=year, month=month, group_id=None)
        emp_entries = [e for e in entries if e.get("employee_id") == emp_id]
        contract_hours = float(emp.get("HRSWEEK") or 0)
        expected_hours = round(contract_hours * working_days / 5, 2) if contract_hours else 0.0
        actual_hours = sum(float(e.get("duration_hours", 0) or 0) for e in emp_entries)
        rows.append({
            "Mitarbeiter": f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(", "),
            "Kürzel": emp.get("SHORTNAME", ""),
            "Vertragl. Std/Woche": contract_hours,
            "Soll-Stunden": expected_hours,
            "Ist-Stunden": round(actual_hours, 2),
            "Differenz": round(actual_hours - expected_hours, 2),
            "Schichten": len(emp_entries),
        })

    filename = f"ueberstunden_{year:04d}-{month:02d}.{fmt}"

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Überstunden {year}-{month:02d}"
            headers = list(rows[0].keys()) if rows else []
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(fill_type="solid", fgColor="1E293B")
                cell.alignment = Alignment(horizontal="center")
            for r_idx, row in enumerate(rows, 2):
                for col, key in enumerate(headers, 1):
                    cell = ws.cell(r_idx, col, row[key])
                    diff = row.get("Differenz", 0)
                    if key == "Differenz" and isinstance(diff, float):
                        cell.font = Font(
                            color="C0392B" if diff < 0 else ("27AE60" if diff > 0 else "000000"),
                            bold=True,
                        )
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), filename
        except ImportError:
            pass  # fall through to CSV

    buf2 = io.StringIO()
    if rows:
        writer = _csv.DictWriter(buf2, fieldnames=list(rows[0].keys()), lineterminator="\r\n")
        writer.writeheader()
        writer.writerows(rows)
    return buf2.getvalue().encode("utf-8-sig"), f"ueberstunden_{year:04d}-{month:02d}.csv"


def _generate_absences_report(year: int, month: int, filters: dict, fmt: str) -> tuple[bytes, str]:
    """Generate an absences statistics report."""
    from sp5lib.db import get_db
    db = get_db()

    group_id = filters.get("group_id")
    employees = db.get_employees(include_hidden=False)
    if group_id is not None:
        member_ids = set(db.get_group_members(group_id))
        employees = [e for e in employees if e["ID"] in member_ids]
    employees.sort(key=lambda x: x.get("POSITION", 0))

    month_start = f"{year:04d}-{month:02d}-01"
    num_days = _calendar.monthrange(year, month)[1]
    month_end = f"{year:04d}-{month:02d}-{num_days:02d}"

    rows = []
    for emp in employees:
        try:
            absences = db.get_absences(
                employee_id=emp["ID"],
                date_from=month_start,
                date_to=month_end,
            )
        except Exception:
            absences = []

        total_days = sum(float(a.get("days", 1)) for a in absences)
        by_type: dict[str, float] = {}
        for a in absences:
            atype = a.get("type_name", a.get("absence_type", "Unbekannt"))
            by_type[atype] = by_type.get(atype, 0) + float(a.get("days", 1))

        rows.append({
            "Mitarbeiter": f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(", "),
            "Kürzel": emp.get("SHORTNAME", ""),
            "Gesamt Abwesenheitstage": total_days,
            "Abwesenheitsarten": "; ".join(f"{k}: {v}" for k, v in by_type.items()),
            "Anzahl Einträge": len(absences),
        })

    filename = f"abwesenheiten_{year:04d}-{month:02d}.{fmt}"

    if fmt == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"Abwesenheiten {year}-{month:02d}"
            headers = list(rows[0].keys()) if rows else []
            for col, h in enumerate(headers, 1):
                cell = ws.cell(1, col, h)
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(fill_type="solid", fgColor="1E293B")
                cell.alignment = Alignment(horizontal="center")
            for r_idx, row in enumerate(rows, 2):
                for col, key in enumerate(headers, 1):
                    ws.cell(r_idx, col, row[key])
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue(), filename
        except ImportError:
            pass

    buf2 = io.StringIO()
    if rows:
        writer = _csv.DictWriter(buf2, fieldnames=list(rows[0].keys()), lineterminator="\r\n")
        writer.writeheader()
        writer.writerows(rows)
    return buf2.getvalue().encode("utf-8-sig"), f"abwesenheiten_{year:04d}-{month:02d}.csv"


def generate_report(report: dict) -> tuple[bytes, str]:
    """Dispatch to the correct report generator. Returns (file_bytes, filename)."""
    report_type = report.get("report_type", "schedule_overview")
    frequency = report.get("frequency", "monthly")
    fmt = report.get("format", "xlsx")
    filters = report.get("filters", {})
    year, month = _get_reference_month(frequency)

    if report_type == "schedule_overview":
        return _generate_schedule_overview(year, month, filters, fmt)
    elif report_type == "overtime":
        return _generate_overtime_report(year, month, filters, fmt)
    elif report_type == "absences":
        return _generate_absences_report(year, month, filters, fmt)
    else:
        raise ValueError(f"Unknown report_type: {report_type}")


# ── Email delivery ─────────────────────────────────────────────────────────────


def send_report_email(report: dict, file_bytes: bytes, filename: str) -> dict:
    """Send the report file via email using the existing email service config."""
    import smtplib
    from email import encoders
    from email.mime.base import MIMEBase
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText

    from sp5lib.email_service import get_config

    cfg = get_config()
    if not cfg.is_configured:
        return {
            "success": False,
            "reason": "SMTP not configured",
        }

    report_type_labels = {
        "schedule_overview": "Dienstplan-Übersicht",
        "overtime": "Überstunden-Auswertung",
        "absences": "Abwesenheits-Statistik",
    }
    report_label = report_type_labels.get(report.get("report_type", ""), report.get("report_type", ""))
    frequency_labels = {"daily": "täglich", "weekly": "wöchentlich", "monthly": "monatlich"}
    freq_label = frequency_labels.get(report.get("frequency", "monthly"), report.get("frequency", ""))

    subject = f"[SP5] {report_label} – {report.get('name', '')}"
    body_plain = (
        f"Automatisch generierter Bericht: {report_label}\n\n"
        f"Name: {report.get('name', '')}\n"
        f"Häufigkeit: {freq_label}\n"
        f"Anhang: {filename}\n\n"
        "---\nOpenSchichtplaner5"
    )
    body_html = f"""\
<!DOCTYPE html><html><head><meta charset="utf-8"></head><body
  style="font-family:sans-serif;color:#1e293b;">
<h2>📊 {report_label}</h2>
<p>Name: <strong>{report.get('name', '')}</strong></p>
<p>Häufigkeit: {freq_label}</p>
<p>Anhang: <code>{filename}</code></p>
<p style="color:#64748b;font-size:12px;">
  Automatisch generiert von OpenSchichtplaner5
</p>
</body></html>"""

    fmt = report.get("format", "xlsx")
    if fmt == "xlsx":
        mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        mime_type = "text/csv"

    sent_to: list[str] = []
    failed: list[str] = []

    for recipient in report.get("recipients", []):
        msg = MIMEMultipart("mixed")
        msg["From"] = cfg.from_addr
        msg["To"] = recipient
        msg["Subject"] = subject

        alt = MIMEMultipart("alternative")
        alt.attach(MIMEText(body_plain, "plain", "utf-8"))
        alt.attach(MIMEText(body_html, "html", "utf-8"))
        msg.attach(alt)

        attachment = MIMEBase(*mime_type.split("/", 1))
        attachment.set_payload(file_bytes)
        encoders.encode_base64(attachment)
        attachment.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(attachment)

        try:
            if cfg.tls_mode == "ssl":
                with smtplib.SMTP_SSL(cfg.host, cfg.port, timeout=15) as srv:
                    if cfg.user:
                        srv.login(cfg.user, cfg.password)
                    srv.send_message(msg)
            else:
                with smtplib.SMTP(cfg.host, cfg.port, timeout=15) as srv:
                    if cfg.tls_mode in ("true", "starttls", "1", "yes"):
                        srv.starttls()
                    if cfg.user:
                        srv.login(cfg.user, cfg.password)
                    srv.send_message(msg)
            sent_to.append(recipient)
        except Exception as exc:
            _logger.warning("Failed to send report email to %s: %s", recipient, exc)
            failed.append(recipient)

    if sent_to:
        return {"success": True, "sent_to": sent_to, "failed": failed, "filename": filename}
    return {"success": False, "reason": "All email sends failed", "failed": failed}


# ── Background scheduler ───────────────────────────────────────────────────────

_scheduler_thread: threading.Thread | None = None
_scheduler_running = False
_scheduler_lock = threading.Lock()
_scheduler_last_run: str | None = None
_scheduler_reports_sent: int = 0


def _run_due_reports() -> int:
    """Check all scheduled reports and send any that are due. Returns count sent."""
    global _scheduler_last_run, _scheduler_reports_sent
    now = _dt.now(UTC)
    reports = _load_reports()
    sent_count = 0

    for report in reports:
        if not report.get("enabled", True):
            continue
        next_run_str = report.get("next_run")
        if not next_run_str:
            continue
        try:
            next_run = _dt.fromisoformat(next_run_str.rstrip("Z")).replace(tzinfo=UTC)
        except (ValueError, AttributeError):
            continue

        if now >= next_run:
            report_id = report.get("id", "unknown")
            try:
                _logger.info("Running scheduled report '%s' (%s)", report.get("name"), report_id)
                file_bytes, filename = generate_report(report)
                result = send_report_email(report, file_bytes, filename)
                _logger.info("Report '%s' send result: %s", report.get("name"), result)
                sent_count += 1
            except Exception as exc:
                _logger.error("Error running scheduled report '%s': %s", report_id, exc)

            # Update next_run regardless of success
            reports_list, idx, _ = _get_report_by_id(report_id)
            reports_list[idx]["next_run"] = _compute_next_run(report.get("frequency", "monthly"), now)
            reports_list[idx]["last_run"] = now.isoformat()
            _save_reports(reports_list)

    _scheduler_last_run = now.isoformat()
    _scheduler_reports_sent += sent_count
    return sent_count


def _scheduler_loop(interval_seconds: int = 300) -> None:
    """Background loop that periodically checks for due reports."""
    global _scheduler_running
    _logger.info("Scheduled reports background scheduler started (interval=%ds)", interval_seconds)
    while _scheduler_running:
        try:
            sent = _run_due_reports()
            if sent > 0:
                _logger.info("Scheduler sent %d reports", sent)
        except Exception as exc:
            _logger.error("Scheduler loop error: %s", exc)
        # Sleep in small increments to allow clean shutdown
        for _ in range(interval_seconds):
            if not _scheduler_running:
                break
            time.sleep(1)
    _logger.info("Scheduled reports background scheduler stopped")


def start_scheduler(interval_seconds: int = 300) -> None:
    """Start the background scheduler thread (idempotent)."""
    global _scheduler_thread, _scheduler_running
    with _scheduler_lock:
        if _scheduler_thread is not None and _scheduler_thread.is_alive():
            return  # Already running
        _scheduler_running = True
        _scheduler_thread = threading.Thread(
            target=_scheduler_loop,
            args=(interval_seconds,),
            daemon=True,
            name="scheduled-reports-scheduler",
        )
        _scheduler_thread.start()


def stop_scheduler() -> None:
    """Stop the background scheduler thread."""
    global _scheduler_running
    _scheduler_running = False


# ── Endpoints ──────────────────────────────────────────────────────────────────


@router.get(
    "",
    summary="List scheduled reports",
    description="Returns all configured scheduled reports. **Required role:** Planer",
)
def list_scheduled_reports(_cur_user: dict = Depends(require_planer)) -> list[dict]:
    """Return all scheduled reports."""
    return _load_reports()


@router.post(
    "",
    summary="Create scheduled report",
    description="Create a new scheduled report. **Required role:** Admin",
    status_code=201,
)
def create_scheduled_report(
    body: ScheduledReportCreate,
    _cur_user: dict = Depends(require_admin),
) -> dict:
    """Create a new scheduled report configuration."""
    reports = _load_reports()
    now = _dt.now(UTC)
    new_report = {
        "id": str(uuid.uuid4()),
        "name": body.name,
        "report_type": body.report_type,
        "frequency": body.frequency,
        "recipients": body.recipients,
        "format": body.format,
        "filters": body.filters,
        "enabled": body.enabled,
        "next_run": _compute_next_run(body.frequency, now),
        "last_run": None,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat(),
    }
    reports.append(new_report)
    _save_reports(reports)
    return new_report


@router.get(
    "/scheduler/status",
    summary="Get scheduler status",
    description="Returns background scheduler status. **Required role:** Admin",
)
def get_scheduler_status(_cur_user: dict = Depends(require_admin)) -> dict:
    """Return the background scheduler status."""
    return {
        "running": _scheduler_running and (_scheduler_thread is not None) and _scheduler_thread.is_alive(),
        "last_run": _scheduler_last_run,
        "reports_sent_total": _scheduler_reports_sent,
        "active_reports": sum(1 for r in _load_reports() if r.get("enabled", True)),
    }


@router.get(
    "/{report_id}",
    summary="Get scheduled report",
    description="Return a single scheduled report by ID. **Required role:** Planer",
)
def get_scheduled_report(
    report_id: str,
    _cur_user: dict = Depends(require_planer),
) -> dict:
    """Return a single scheduled report."""
    _, _, report = _get_report_by_id(report_id)
    return report


@router.put(
    "/{report_id}",
    summary="Update scheduled report",
    description="Update an existing scheduled report. **Required role:** Admin",
)
def update_scheduled_report(
    report_id: str,
    body: ScheduledReportUpdate,
    _cur_user: dict = Depends(require_admin),
) -> dict:
    """Update an existing scheduled report configuration."""
    reports, idx, existing = _get_report_by_id(report_id)
    update_data = body.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        existing[key] = value
    # Recompute next_run if frequency changed
    if "frequency" in update_data:
        existing["next_run"] = _compute_next_run(existing["frequency"])
    existing["updated_at"] = _dt.now(UTC).isoformat()
    reports[idx] = existing
    _save_reports(reports)
    return existing


@router.delete(
    "/{report_id}",
    summary="Delete scheduled report",
    description="Delete a scheduled report permanently. **Required role:** Admin",
    status_code=204,
)
def delete_scheduled_report(
    report_id: str,
    _cur_user: dict = Depends(require_admin),
) -> None:
    """Delete a scheduled report."""
    reports, idx, _ = _get_report_by_id(report_id)
    reports.pop(idx)
    _save_reports(reports)


@router.post(
    "/{report_id}/run",
    summary="Manually trigger a scheduled report",
    description=(
        "Generate the report immediately and send it via email.\n\n"
        "If SMTP is not configured, returns an error.\n\n"
        "**Required role:** Admin"
    ),
)
def run_scheduled_report(
    report_id: str,
    _cur_user: dict = Depends(require_admin),
) -> dict:
    """Manually trigger a scheduled report: generate and email it."""
    reports, idx, report = _get_report_by_id(report_id)
    try:
        file_bytes, filename = generate_report(report)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Report generation failed: {exc}",
        ) from exc

    result = send_report_email(report, file_bytes, filename)

    # Update last_run timestamp
    reports[idx]["last_run"] = _dt.now(UTC).isoformat()
    _save_reports(reports)

    return result
