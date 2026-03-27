"""Tests for Excel XLSX export endpoints (Q051)."""
import io
import os
import sys

import pytest

_BACKEND_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _BACKEND_DIR not in sys.path:
    sys.path.insert(0, _BACKEND_DIR)

import secrets

from fastapi.testclient import TestClient


@pytest.fixture(scope="module")
def client():
    """Create a test client with planer auth."""
    from api.main import _sessions, app

    tok = secrets.token_hex(20)
    _sessions[tok] = {
        "ID": 999,
        "NAME": "test_planer_xlsx",
        "role": "Planer",
        "ADMIN": False,
        "RIGHTS": 2,
    }
    c = TestClient(app)
    c.headers["X-Auth-Token"] = tok
    yield c
    _sessions.pop(tok, None)


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestScheduleXlsx:
    def test_schedule_xlsx_returns_xlsx(self, client):
        resp = client.get("/api/export/schedule?month=2026-03&format=xlsx")
        assert resp.status_code == 200
        assert XLSX_MIME in resp.headers.get("content-type", "")
        assert "dienstplan_2026-03.xlsx" in resp.headers.get("content-disposition", "")

    def test_schedule_xlsx_valid_workbook(self, client):
        resp = client.get("/api/export/schedule?month=2026-03&format=xlsx")
        assert resp.status_code == 200
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.freeze_panes == "C2"
        assert ws.cell(1, 1).value == "Mitarbeiter"
        assert ws.cell(1, 2).value == "Kürzel"
        assert ws.cell(1, 1).font.bold is True


class TestStatisticsXlsx:
    def test_statistics_xlsx_returns_xlsx(self, client):
        resp = client.get("/api/export/statistics?year=2026&format=xlsx")
        assert resp.status_code == 200
        assert XLSX_MIME in resp.headers.get("content-type", "")
        assert "statistiken_2026.xlsx" in resp.headers.get("content-disposition", "")

    def test_statistics_xlsx_has_two_sheets(self, client):
        resp = client.get("/api/export/statistics?year=2026&format=xlsx")
        assert resp.status_code == 200
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert len(wb.sheetnames) == 2
        assert "Jahresübersicht" in wb.sheetnames[0]
        assert "Monatsdetail" in wb.sheetnames[1]

    def test_statistics_xlsx_freeze_panes(self, client):
        resp = client.get("/api/export/statistics?year=2026&format=xlsx")
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        for ws in wb.worksheets:
            assert ws.freeze_panes == "A2"

    def test_statistics_xlsx_headers(self, client):
        resp = client.get("/api/export/statistics?year=2026&format=xlsx")
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.worksheets[0]
        assert ws.cell(1, 1).value == "Mitarbeiter"
        assert ws.cell(1, 3).value == "Soll (h)"
        assert ws.cell(1, 1).font.bold is True


class TestEmployeesXlsx:
    def test_employees_xlsx_returns_xlsx(self, client):
        resp = client.get("/api/export/employees?format=xlsx")
        assert resp.status_code == 200
        assert XLSX_MIME in resp.headers.get("content-type", "")
        assert "mitarbeiter_" in resp.headers.get("content-disposition", "")

    def test_employees_xlsx_structure(self, client):
        resp = client.get("/api/export/employees?format=xlsx")
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        if ws.cell(1, 1).value is not None:
            assert ws.freeze_panes == "A2"
            assert ws.cell(1, 1).value == "ID"
            assert ws.cell(1, 1).font.bold is True


class TestAbsencesXlsx:
    def test_absences_xlsx_returns_xlsx(self, client):
        resp = client.get("/api/export/absences?year=2026&format=xlsx")
        assert resp.status_code == 200
        assert XLSX_MIME in resp.headers.get("content-type", "")
        assert "abwesenheiten_2026.xlsx" in resp.headers.get("content-disposition", "")

    def test_absences_xlsx_valid_workbook(self, client):
        resp = client.get("/api/export/absences?year=2026&format=xlsx")
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.active
        assert ws.title.startswith("Abwesenheiten")


class TestXlsxFormatConsistency:
    """Ensure all xlsx exports return valid xlsx files."""

    @pytest.mark.parametrize(
        "url,filename_part",
        [
            ("/api/export/schedule?month=2026-03&format=xlsx", "dienstplan"),
            ("/api/export/employees?format=xlsx", "mitarbeiter"),
            ("/api/export/absences?year=2026&format=xlsx", "abwesenheiten"),
            ("/api/export/statistics?year=2026&format=xlsx", "statistiken"),
        ],
    )
    def test_valid_xlsx_file(self, client, url, filename_part):
        resp = client.get(url)
        assert resp.status_code == 200
        assert XLSX_MIME in resp.headers.get("content-type", "")
        assert filename_part in resp.headers.get("content-disposition", "")
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        assert len(wb.sheetnames) >= 1

    @pytest.mark.parametrize(
        "url",
        [
            "/api/export/schedule?month=2026-03&format=xlsx",
            "/api/export/statistics?year=2026&format=xlsx",
        ],
    )
    def test_freeze_panes_always_set(self, client, url):
        resp = client.get(url)
        assert resp.status_code == 200
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(resp.content))
        ws = wb.worksheets[0]
        assert ws.freeze_panes is not None
