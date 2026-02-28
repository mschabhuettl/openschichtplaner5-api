"""Reports, statistics, zeitkonto, export, import, analysis router."""
import io
import csv
import html as _html
import calendar as _calendar
from datetime import datetime as _dt, date
from fastapi import APIRouter, HTTPException, Query, Depends, Request, UploadFile, File
from fastapi.responses import Response as _Response
from pydantic import BaseModel, Field
from typing import Optional
from ..dependencies import (
    get_db, require_planer, require_auth, _sanitize_500, limiter,
)

router = APIRouter()


def _xlsx_response(content: bytes, filename: str) -> _Response:
    return _Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Monthly statistics ───────────────────────────────────────
@router.get(
    "/api/statistics",
    tags=["Statistics"],
    summary="Monthly statistics",
    description=(
        "Return per-employee statistics for a given month.\n\n"
        "Each row contains target hours, actual hours, overtime, vacation days, and sick days. "
        "Optionally filter by `group_id`. Defaults to the current year/month if not specified.\n\n"
        "**Required role:** Leser (read-only)"
    ),
)
def get_statistics(
    year: Optional[int] = Query(None, description="Year (YYYY), defaults to current year"),
    month: Optional[int] = Query(None, description="Month (1-12), defaults to current month"),
    group_id: Optional[int] = Query(None, description="Filter by group ID"),
):
    from datetime import date as _date
    if year is None:
        year = _date.today().year
    if month is None:
        month = _date.today().month
    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="Ungültiger Monat: muss zwischen 1 und 12 liegen")
    return get_db().get_statistics(year, month, group_id=group_id)


# ── Year Summary (Jahresrückblick) ────────────────────────────
@router.get(
    "/api/statistics/year-summary",
    tags=["Statistics"],
    summary="Year summary (Jahresrückblick)",
    description=(
        "Return aggregated statistics for all 12 months of a year (Jahresrückblick).\n\n"
        "Contains per-employee totals for the full year plus month-by-month breakdown. "
        "Optionally filter by `group_id`.\n\n"
        "**Required role:** Leser"
    ),
)
def get_year_summary(
    year: Optional[int] = Query(None, description="Year (YYYY), defaults to current year"),
    group_id: Optional[int] = Query(None, description="Filter by group ID"),
):
    """Return aggregated statistics for all 12 months of a year (Jahresrückblick)."""
    from datetime import date as _date
    if year is None:
        year = _date.today().year
    db = get_db()

    # Collect stats for each month — single pass to avoid duplicate DB calls
    monthly = []
    emp_totals: dict = {}
    all_monthly_rows: list = []
    for m in range(1, 13):
        rows = db.get_statistics(year, m, group_id=group_id)
        all_monthly_rows.append((m, rows))
        total_actual = sum(r.get("actual_hours", 0) or 0 for r in rows)
        total_target = sum(r.get("target_hours", 0) or 0 for r in rows)
        total_absences = sum(r.get("absence_days", 0) or 0 for r in rows)
        total_vacation = sum(r.get("vacation_used", 0) or 0 for r in rows)
        total_sick = sum(r.get("sick_days", 0) or 0 for r in rows)
        total_shifts = sum(r.get("shifts_count", 0) or 0 for r in rows)
        employee_count = len(rows)
        monthly.append({
            "month": m,
            "actual_hours": round(total_actual, 1),
            "target_hours": round(total_target, 1),
            "absence_days": total_absences,
            "vacation_days": total_vacation,
            "sick_days": total_sick,
            "shifts_count": total_shifts,
            "employee_count": employee_count,
            "overtime": round(total_actual - total_target, 1),
        })

    # Per-employee year totals (re-use already-fetched rows)
    for m, rows in all_monthly_rows:
        for r in rows:
            eid = r.get("employee_id")
            if eid not in emp_totals:
                emp_totals[eid] = {
                    "employee_id": eid,
                    "name": r.get("employee_name", r.get("name", "")),
                    "group": r.get("group_name", r.get("group", "")),
                    "actual_hours": 0.0,
                    "target_hours": 0.0,
                    "absence_days": 0,
                    "vacation_days": 0,
                    "sick_days": 0,
                    "shifts_count": 0,
                    "monthly_hours": [0.0] * 12,
                }
            emp_totals[eid]["actual_hours"] += r.get("actual_hours", 0) or 0
            emp_totals[eid]["target_hours"] += r.get("target_hours", 0) or 0
            emp_totals[eid]["absence_days"] += r.get("absence_days", 0) or 0
            emp_totals[eid]["vacation_days"] += r.get("vacation_used", 0) or 0
            emp_totals[eid]["sick_days"] += r.get("sick_days", 0) or 0
            emp_totals[eid]["shifts_count"] += r.get("shifts_count", 0) or 0
            emp_totals[eid]["monthly_hours"][m - 1] = round(r.get("actual_hours", 0) or 0, 1)

    employees = sorted(emp_totals.values(), key=lambda x: x["actual_hours"], reverse=True)
    for e in employees:
        e["actual_hours"] = round(e["actual_hours"], 1)
        e["target_hours"] = round(e["target_hours"], 1)
        e["overtime"] = round(e["actual_hours"] - e["target_hours"], 1)

    # Year totals
    year_totals = {
        "actual_hours": round(sum(m["actual_hours"] for m in monthly), 1),
        "target_hours": round(sum(m["target_hours"] for m in monthly), 1),
        "absence_days": sum(m["absence_days"] for m in monthly),
        "vacation_days": sum(m["vacation_days"] for m in monthly),
        "sick_days": sum(m["sick_days"] for m in monthly),
        "shifts_count": sum(m["shifts_count"] for m in monthly),
    }
    year_totals["overtime"] = round(year_totals["actual_hours"] - year_totals["target_hours"], 1)

    return {
        "year": year,
        "monthly": monthly,
        "employees": employees,
        "totals": year_totals,
    }


# ── Employee detailed statistics ─────────────────────────────
@router.get(
    "/api/statistics/employee/{emp_id}",
    tags=["Statistics"],
    summary="Employee statistics",
    description=(
        "Return detailed statistics for a single employee.\n\n"
        "Without `month`: returns 12-month yearly breakdown. "
        "With `month`: returns daily schedule details for that month.\n\n"
        "**Required role:** Leser"
    ),
)
def get_employee_statistics(
    emp_id: int,
    year: Optional[int] = Query(None, description="Year (YYYY), defaults to current year"),
    month: Optional[int] = Query(None, description="Month (1-12); if omitted returns full year overview"),
):
    """
    Return detailed statistics for a single employee.
    Without month: returns 12-month yearly breakdown.
    With month: returns stats for that specific month only.
    """
    from datetime import date as _date
    if year is None:
        year = _date.today().year
    db = get_db()
    emp = db.get_employee(emp_id)
    if emp is None:
        raise HTTPException(status_code=404, detail=f"Employee {emp_id} not found")
    if month is not None:
        if not (1 <= month <= 12):
            raise HTTPException(status_code=400, detail="Ungültiger Monat: muss zwischen 1 und 12 liegen")
        return db.get_employee_stats_month(emp_id, year, month)
    return db.get_employee_stats_year(emp_id, year)


# ── Sickness / Krankenstand statistics ───────────────────────
@router.get(
    "/api/statistics/sickness",
    tags=["Statistics"],
    summary="Sickness statistics (Krankenstand)",
    description=(
        "Return sickness (Krankenstand) statistics for a given year.\n\n"
        "Response contains:\n"
        "- `per_employee`: sick days, episodes, Bradford factor per employee\n"
        "- `per_month`: 12-element list with monthly sick-day totals\n"
        "- `per_weekday`: 7-element list with per-weekday sick-day totals\n"
        "- `totals`: total_sick_days, affected_employees, total_employees\n\n"
        "**Required role:** Leser"
    ),
)
def get_sickness_statistics(
    year: Optional[int] = Query(None, description="Year (YYYY), defaults to current year"),
):
    """Return sickness (Krankenstand) statistics for a given year.

    Response contains:
    - per_employee: list of {employee_id, name, sick_days, sick_episodes, bradford_factor}
    - per_month: 12-element list {month, sick_days}
    - per_weekday: 7-element list {weekday, weekday_name, sick_days}
    - totals: total_sick_days, affected_employees, total_employees
    """
    from datetime import date as _date
    if year is None:
        year = _date.today().year
    return get_db().get_sickness_statistics(year)


# ── Shift statistics ──────────────────────────────────────────
@router.get(
    "/api/statistics/shifts",
    tags=["Statistics"],
    summary="Shift statistics (trend)",
    description=(
        "Return shift-centric statistics over a rolling window of months.\n\n"
        "Response contains:\n"
        "- `periods`: list of {year, month, label} for the trend window\n"
        "- `shift_usage`: per-shift monthly counts and totals\n"
        "- `employee_distribution`: per-employee shift counts by category\n"
        "- `category_totals`: global counts by category (Früh/Spät/Nacht/Sonstige)\n\n"
        "**Required role:** Leser"
    ),
)
def get_shift_statistics(
    year: int = Query(..., description="Year (YYYY)"),
    months: int = Query(6, ge=1, le=24, description="Number of months for trend"),
    group_id: Optional[int] = Query(None),
):
    """
    Return shift-centric statistics:
    - periods: list of {year, month, label} for trend window
    - shift_usage: per shift, monthly counts + total
    - employee_distribution: per employee, counts by shift category
    - category_totals: global counts by category (Früh/Spät/Nacht/Sonstige)
    """
    from datetime import date as _date
    from collections import defaultdict

    db = get_db()
    shifts_map = {s['ID']: s for s in db.get_shifts(include_hidden=True)}

    if group_id is not None:
        member_ids = set(db.get_group_members(group_id))
    else:
        member_ids = None

    employees = {e['ID']: e for e in db.get_employees(include_hidden=False)}

    # Build list of (year, month) for trend window (most recent `months` months)
    today = _date.today()
    end_year, end_month = today.year, today.month
    periods = []
    y, m = end_year, end_month
    for _ in range(months):
        periods.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    periods.reverse()
    period_set = set(periods)

    shift_month_counts: dict = defaultdict(lambda: defaultdict(int))
    emp_shift_counts: dict = defaultdict(lambda: defaultdict(int))

    for r in db._read('MASHI'):
        d = r.get('DATE', '')
        if not d or len(d) < 7:
            continue
        try:
            ry = int(d[0:4])
            rm = int(d[5:7])
        except (ValueError, IndexError):
            continue
        if (ry, rm) not in period_set:
            continue
        eid = r.get('EMPLOYEEID')
        if member_ids is not None and eid not in member_ids:
            continue
        sid = r.get('SHIFTID')
        if not sid:
            continue
        shift_month_counts[sid][(ry, rm)] += 1
        emp_shift_counts[eid][sid] += 1

    def categorize_shift(s: dict) -> str:
        start = s.get('FROM0') or ''
        if start and isinstance(start, str) and ':' in start:
            try:
                hour = int(start.split(':')[0])
                if 4 <= hour < 11:
                    return 'Früh'
                elif 11 <= hour < 18:
                    return 'Spät'
                elif hour >= 18 or hour < 4:
                    return 'Nacht'
            except ValueError:
                pass
        name = (s.get('NAME', '') or '').upper()
        short = (s.get('SHORTNAME', '') or '').upper()
        if 'FRÜH' in name or 'FRUH' in name or short in ('F', 'FR'):
            return 'Früh'
        if 'SPÄT' in name or 'SPAT' in name or short in ('S', 'SP'):
            return 'Spät'
        if 'NACHT' in name or 'NIGHT' in name or short in ('N', 'NA'):
            return 'Nacht'
        return 'Sonstige'

    shift_usage = []
    for sid, month_map in shift_month_counts.items():
        s = shifts_map.get(sid, {})
        monthly = [{'year': ry, 'month': rm, 'count': month_map.get((ry, rm), 0)} for (ry, rm) in periods]
        total = sum(month_map.values())
        shift_usage.append({
            'shift_id': sid,
            'name': s.get('NAME', str(sid)),
            'short': s.get('SHORTNAME', ''),
            'color_bk': s.get('COLORBK', None),
            'color_text': s.get('COLORTEXT', None),
            'category': categorize_shift(s),
            'monthly_counts': monthly,
            'total': total,
        })
    shift_usage.sort(key=lambda x: -x['total'])

    cat_counts_global: dict = defaultdict(int)
    emp_distribution = []
    for eid, shift_counts in emp_shift_counts.items():
        emp = employees.get(eid)
        if not emp:
            continue
        by_category: dict = defaultdict(int)
        for sid, cnt in shift_counts.items():
            cat = categorize_shift(shifts_map.get(sid, {}))
            by_category[cat] += cnt
            cat_counts_global[cat] += cnt
        emp_distribution.append({
            'employee_id': eid,
            'name': (emp.get('LASTNAME', '') + ' ' + emp.get('FIRSTNAME', '')).strip(),
            'short': emp.get('SHORTNAME', ''),
            'total_shifts': sum(shift_counts.values()),
            'by_category': dict(by_category),
        })
    emp_distribution.sort(key=lambda x: -x['total_shifts'])

    month_names_short = ['', 'Jan', 'Feb', 'Mär', 'Apr', 'Mai', 'Jun',
                         'Jul', 'Aug', 'Sep', 'Okt', 'Nov', 'Dez']
    period_labels = [{'year': ry, 'month': rm, 'label': f"{month_names_short[rm]} {ry}"} for (ry, rm) in periods]

    return {
        'periods': period_labels,
        'shift_usage': shift_usage,
        'employee_distribution': emp_distribution,
        'category_totals': dict(cat_counts_global),
    }


# ── Export endpoints ─────────────────────────────────────────

from fastapi.responses import Response as _Response  # noqa: E402


def _int_to_rgb(color_int: int) -> str:
    """Convert BGR int to #RRGGBB hex."""
    b = (color_int >> 16) & 0xFF
    g = (color_int >> 8) & 0xFF
    r = color_int & 0xFF
    return f"#{r:02X}{g:02X}{b:02X}"


def _csv_response(rows: list, filename: str) -> _Response:
    buf = io.StringIO()
    if rows:
        writer = csv.DictWriter(buf, fieldnames=rows[0].keys(), lineterminator='\r\n')
        writer.writeheader()
        writer.writerows(rows)
    content = buf.getvalue()
    return _Response(
        content=content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get(
    "/api/export/schedule",
    tags=["Export"],
    summary="Export schedule",
    description=(
        "Export the monthly schedule as CSV, HTML, or XLSX.\n\n"
        "- `month`: required, format `YYYY-MM`\n"
        "- `group_id`: optional filter\n"
        "- `format`: `csv` (default), `html`, or `xlsx`\n\n"
        "Returns a file download with appropriate Content-Disposition header.\n\n"
        "**Required role:** Planer"
    ),
    responses={
        200: {"description": "File download (CSV/HTML/XLSX)"},
        400: {"description": "Invalid month format or format type"},
        401: {"description": "Not authenticated"},
        429: {"description": "Rate limit exceeded"},
    },
)
@limiter.limit("10/minute")
def export_schedule(
    request: Request,
    month: str = Query(..., description="Month in YYYY-MM format"),
    group_id: Optional[int] = Query(None),
    format: str = Query("csv", description="csv, html, or xlsx"),
    _cur_user: dict = Depends(require_planer),
):
    try:
        dt = _dt.strptime(month, "%Y-%m")
        year, mon = dt.year, dt.month
    except ValueError:
        raise HTTPException(status_code=400, detail="month must be YYYY-MM")

    db = get_db()
    entries = db.get_schedule(year=year, month=mon, group_id=group_id)
    employees = db.get_employees(include_hidden=False)
    if group_id is not None:
        member_ids = set(db.get_group_members(group_id))
        employees = [e for e in employees if e['ID'] in member_ids]
    employees.sort(key=lambda x: x.get('POSITION', 0))

    # Build lookup: (emp_id, date) -> entry
    entry_map: dict = {}
    for e in entries:
        key = (e['employee_id'], e['date'])
        entry_map[key] = e

    num_days = _calendar.monthrange(year, mon)[1]
    days = [f"{year:04d}-{mon:02d}-{d:02d}" for d in range(1, num_days + 1)]

    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: F401
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl nicht installiert.")
        _month_names_de = ["Januar", "Februar", "März", "April", "Mai", "Juni",
                           "Juli", "August", "September", "Oktober", "November", "Dezember"]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{_month_names_de[mon-1]} {year}"
        # Header row
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(bold=True, color="FFFFFF", size=9)
        header_fill = PatternFill(fill_type="solid", fgColor="1E293B")
        # Column 1: Name, col 2: Kürzel, then days
        ws.cell(1, 1, "Mitarbeiter").font = header_font
        ws.cell(1, 1).fill = header_fill
        ws.cell(1, 1).alignment = Alignment(horizontal="left")
        ws.cell(1, 1).border = border
        ws.column_dimensions['A'].width = 22
        ws.cell(1, 2, "Kürzel").font = header_font
        ws.cell(1, 2).fill = header_fill
        ws.cell(1, 2).alignment = Alignment(horizontal="center")
        ws.cell(1, 2).border = border
        ws.column_dimensions['B'].width = 6
        for d in range(1, num_days + 1):
            col = d + 2
            wd = _dt(year, mon, d).weekday()
            wd_abbr = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"][wd]
            cell = ws.cell(1, col, f"{d}\n{wd_abbr}")
            cell.font = header_font
            is_weekend = wd >= 5
            cell.fill = PatternFill(fill_type="solid", fgColor="475569" if is_weekend else "1E293B")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = 4.5
        ws.row_dimensions[1].height = 28
        # Data rows
        for r_idx, emp in enumerate(employees, start=2):
            emp_name = f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(', ')
            short = emp.get('SHORTNAME', '')
            # Employee name cell
            cbklabel = emp.get('CBKLABEL', 16777215)
            cbklabel_hex = emp.get('CBKLABEL_HEX', '#f8fafc')
            cfglabel_hex = emp.get('CFGLABEL_HEX', '#000000')
            emp_bg = cbklabel_hex.lstrip('#') if (cbklabel and cbklabel != 16777215 and cbklabel != 0) else "F8FAFC"
            emp_fg = cfglabel_hex.lstrip('#') if (cbklabel and cbklabel != 16777215 and cbklabel != 0) else "1E293B"
            name_cell = ws.cell(r_idx, 1, emp_name)
            name_cell.font = Font(bold=bool(emp.get('BOLD')), color=emp_fg, size=9)
            name_cell.fill = PatternFill(fill_type="solid", fgColor=emp_bg)
            name_cell.border = border
            short_cell = ws.cell(r_idx, 2, short)
            short_cell.font = Font(color="64748B", size=9)
            short_cell.fill = PatternFill(fill_type="solid", fgColor="F8FAFC")
            short_cell.alignment = Alignment(horizontal="center")
            short_cell.border = border
            for d in range(1, num_days + 1):
                col = d + 2
                date_str = f"{year:04d}-{mon:02d}-{d:02d}"
                wd = _dt(year, mon, d).weekday()
                is_weekend = wd >= 5
                e = entry_map.get((emp['ID'], date_str))
                cell = ws.cell(r_idx, col)
                if e:
                    bg = e.get('color_bk', '#4A90D9').lstrip('#')
                    fg = e.get('color_text', '#FFFFFF').lstrip('#')
                    cell.value = e.get('display_name', '')
                    cell.font = Font(bold=True, color=fg, size=8)
                    cell.fill = PatternFill(fill_type="solid", fgColor=bg)
                else:
                    cell.fill = PatternFill(fill_type="solid", fgColor="EBEBEB" if is_weekend else "FFFFFF")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            ws.row_dimensions[r_idx].height = 14
        buf = io.BytesIO()
        wb.save(buf)
        return _xlsx_response(buf.getvalue(), f"dienstplan_{month}.xlsx")

    if format == "csv":
        rows = []
        for emp in employees:
            row: dict = {
                "Mitarbeiter": f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(', '),
                "Kürzel": emp.get('SHORTNAME', ''),
            }
            for date in days:
                day_num = int(date.split('-')[2])
                e = entry_map.get((emp['ID'], date))
                row[str(day_num)] = e['display_name'] if e else ''
            rows.append(row)
        return _csv_response(rows, f"dienstplan_{month}.csv")
    else:
        # HTML export
        _month_names_de = ["Januar", "Februar", "März", "April", "Mai", "Juni",
                           "Juli", "August", "September", "Oktober", "November", "Dezember"]
        month_name = f"{_month_names_de[mon - 1]} {year}"
        day_headers = ""
        for d in range(1, num_days + 1):
            wd = _dt(year, mon, d).weekday()
            wd_abbr = ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"][wd]
            is_weekend = wd >= 5
            cls = "weekend" if is_weekend else ""
            day_headers += f'<th class="day-header {cls}">{d}<br><span style="font-weight:normal;font-size:9px">{wd_abbr}</span></th>'

        # Build shift legend
        shifts_all = db.get_shifts(include_hidden=False)
        db.get_leave_types(include_hidden=False)
        legend_html = '<div class="no-print" style="margin-top:12px;display:flex;flex-wrap:wrap;gap:6px;align-items:center"><strong style="font-size:11px;color:#334155">Legende:</strong>'
        for s in shifts_all:
            bg = s.get('COLORBK_HEX', '#fff')
            fg = s.get('COLORTEXT_HEX', '#000')
            name = s.get('NAME', '')
            short = s.get('SHORTNAME', '')
            legend_html += f'<span style="background:{bg};color:{fg};padding:2px 6px;border:1px solid #ccc;border-radius:3px;font-size:10px;font-weight:bold" title="{_html.escape(name)}">{_html.escape(short)}</span>'
        legend_html += '</div>'

        rows_html = ""
        for emp in employees:
            emp_name = f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(', ')
            short = emp.get('SHORTNAME', '')
            # Use employee's label color (CBKLABEL) if not white/default
            cbklabel = emp.get('CBKLABEL', 16777215)
            cbklabel_hex = emp.get('CBKLABEL_HEX', '#f8fafc')
            cfglabel_hex = emp.get('CFGLABEL_HEX', '#000000')
            bold_style = 'font-weight:bold;' if emp.get('BOLD') else ''
            if cbklabel and cbklabel != 16777215 and cbklabel != 0:
                emp_style = f'background:{cbklabel_hex};color:{cfglabel_hex};{bold_style}'
            else:
                emp_style = f'background:#f8fafc;{bold_style}'
            rows_html += f'<tr><td class="emp-name" style="{emp_style}">{_html.escape(emp_name)}</td><td class="emp-short">{_html.escape(short)}</td>'
            for date in days:
                wd = _dt(year, mon, int(date.split('-')[2])).weekday()
                is_weekend = wd >= 5
                e = entry_map.get((emp['ID'], date))
                if e:
                    bg = e.get('color_bk', '#4A90D9')
                    fg = e.get('color_text', '#FFFFFF')
                    display = e.get('display_name', '')
                    rows_html += f'<td class="day-cell" style="background:{bg};color:{fg}"><span title="{_html.escape(str(e.get("shift_name", e.get("leave_name", display))))}">{_html.escape(str(display))}</span></td>'
                else:
                    weekend_style = 'background:#f0f0f0;' if is_weekend else ''
                    rows_html += f'<td class="day-cell" style="{weekend_style}"></td>'
            rows_html += '</tr>\n'

        html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<title>Dienstplan {month_name}</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; }}
  h1 {{ font-size: 16px; margin-bottom: 8px; color: #1e293b; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ border: 1px solid #d1d5db; padding: 3px 4px; white-space: nowrap; }}
  th {{ background: #1e293b; color: white; text-align: center; font-size: 10px; }}
  .emp-name {{ background: #f8fafc; font-weight: bold; min-width: 120px; }}
  .emp-short {{ background: #f8fafc; text-align: center; min-width: 36px; color: #64748b; }}
  .day-header {{ min-width: 28px; }}
  .day-cell {{ text-align: center; font-size: 10px; font-weight: bold; }}
  .weekend {{ background: #475569 !important; }}
  @media print {{
    body {{ margin: 5mm; }}
    .no-print {{ display: none; }}
  }}
</style>
</head>
<body>
<h1>📅 Dienstplan — {month_name}</h1>
<p class="no-print" style="color:#64748b;font-size:11px">Gedruckt am {_dt.now().strftime("%d.%m.%Y %H:%M")}</p>
{legend_html}
<table>
<thead>
<tr>
  <th style="text-align:left;min-width:130px">Mitarbeiter</th>
  <th style="min-width:36px">Kürzel</th>
  {day_headers}
</tr>
</thead>
<tbody>
{rows_html}
</tbody>
</table>
</body>
</html>"""
        return _Response(
            content=html,
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="dienstplan_{month}.html"'},
        )


@router.get(
    "/api/export/statistics",
    tags=["Export"],
    summary="Export yearly statistics",
    description=(
        "Export yearly statistics (all 12 months) as CSV or HTML.\n\n"
        "Includes per-employee target hours, actual hours, overtime, vacation days, and sick days "
        "plus a summary row per employee across the year.\n\n"
        "**Required role:** Planer"
    ),
    responses={
        200: {"description": "File download (CSV/HTML)"},
        401: {"description": "Not authenticated"},
        429: {"description": "Rate limit exceeded"},
    },
)
@limiter.limit("10/minute")
def export_statistics(
    request: Request,
    year: int = Query(...),
    group_id: Optional[int] = Query(None),
    format: str = Query("csv", description="csv or html"),
    _cur_user: dict = Depends(require_planer),
):
    db = get_db()
    rows_data = []
    for mon in range(1, 13):
        month_stats = db.get_statistics(year=year, month=mon, group_id=group_id)
        for s in month_stats:
            rows_data.append({
                "Monat": mon,
                "Mitarbeiter": s['employee_name'],
                "Kürzel": s['employee_short'],
                "Soll (h)": s['target_hours'],
                "Ist (h)": s['actual_hours'],
                "Überstunden (h)": s['overtime_hours'],
                "Abwesenheitstage": s['absence_days'],
                "Urlaubstage": s['vacation_used'],
            })

    # Also build a summary per employee (sum over year)
    from collections import defaultdict
    summary: dict = defaultdict(lambda: {
        "Mitarbeiter": "", "Kürzel": "",
        "Soll (h)": 0.0, "Ist (h)": 0.0, "Überstunden (h)": 0.0,
        "Abwesenheitstage": 0, "Urlaubstage": 0,
    })
    for r in rows_data:
        k = r["Mitarbeiter"]
        summary[k]["Mitarbeiter"] = r["Mitarbeiter"]
        summary[k]["Kürzel"] = r["Kürzel"]
        summary[k]["Soll (h)"] += r["Soll (h)"]
        summary[k]["Ist (h)"] += r["Ist (h)"]
        summary[k]["Überstunden (h)"] += r["Überstunden (h)"]
        summary[k]["Abwesenheitstage"] += r["Abwesenheitstage"]
        summary[k]["Urlaubstage"] += r["Urlaubstage"]

    if format == "csv":
        return _csv_response(rows_data, f"statistiken_{year}.csv")
    else:
        MONTHS_DE = ["", "Januar", "Februar", "März", "April", "Mai", "Juni",
                     "Juli", "August", "September", "Oktober", "November", "Dezember"]

        # Build summary table rows
        summary_rows = ""
        for s in summary.values():
            ot = s["Überstunden (h)"]
            ot_color = "#16a34a" if ot >= 0 else "#dc2626"
            summary_rows += (
                f'<tr>'
                f'<td class="name">{_html.escape(str(s["Mitarbeiter"]))}</td>'
                f'<td class="center">{_html.escape(str(s["Kürzel"]))}</td>'
                f'<td class="num">{s["Soll (h)"]:.1f}</td>'
                f'<td class="num">{s["Ist (h)"]:.1f}</td>'
                f'<td class="num" style="color:{ot_color};font-weight:bold">{"+" if ot>=0 else ""}{ot:.1f}</td>'
                f'<td class="num">{s["Abwesenheitstage"]}</td>'
                f'<td class="num">{s["Urlaubstage"]}</td>'
                f'</tr>\n'
            )

        # Build monthly detail rows
        detail_rows = ""
        for r in rows_data:
            ot = r["Überstunden (h)"]
            ot_color = "#16a34a" if ot >= 0 else "#dc2626"
            detail_rows += (
                f'<tr>'
                f'<td class="center">{_html.escape(str(MONTHS_DE[r["Monat"]]))}</td>'
                f'<td class="name">{_html.escape(str(r["Mitarbeiter"]))}</td>'
                f'<td class="center">{_html.escape(str(r["Kürzel"]))}</td>'
                f'<td class="num">{r["Soll (h)"]:.1f}</td>'
                f'<td class="num">{r["Ist (h)"]:.1f}</td>'
                f'<td class="num" style="color:{ot_color};font-weight:bold">{"+" if ot>=0 else ""}{ot:.1f}</td>'
                f'<td class="num">{r["Abwesenheitstage"]}</td>'
                f'<td class="num">{r["Urlaubstage"]}</td>'
                f'</tr>\n'
            )

        html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<title>Statistiken {year}</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; }}
  h1 {{ font-size: 16px; color: #1e293b; margin-bottom: 4px; }}
  h2 {{ font-size: 13px; color: #334155; margin: 18px 0 6px; }}
  table {{ border-collapse: collapse; width: 100%; margin-bottom: 24px; }}
  th, td {{ border: 1px solid #d1d5db; padding: 4px 6px; }}
  th {{ background: #1e293b; color: white; text-align: center; }}
  .name {{ font-weight: bold; }}
  .center {{ text-align: center; }}
  .num {{ text-align: right; }}
  tr:nth-child(even) {{ background: #f8fafc; }}
  @media print {{ body {{ margin: 5mm; }} }}
</style>
</head>
<body>
<h1>📈 Statistiken — {year}</h1>
<p style="color:#64748b;font-size:11px">Erstellt am {_dt.now().strftime("%d.%m.%Y %H:%M")}</p>

<h2>Jahresübersicht (gesamt)</h2>
<table>
<thead>
<tr>
  <th style="text-align:left">Mitarbeiter</th>
  <th>Kürzel</th>
  <th>Soll (h)</th>
  <th>Ist (h)</th>
  <th>Überstunden</th>
  <th>Abwesenheiten</th>
  <th>Urlaub</th>
</tr>
</thead>
<tbody>
{summary_rows}
</tbody>
</table>

<h2>Monatsdetail</h2>
<table>
<thead>
<tr>
  <th>Monat</th>
  <th style="text-align:left">Mitarbeiter</th>
  <th>Kürzel</th>
  <th>Soll (h)</th>
  <th>Ist (h)</th>
  <th>Überstunden</th>
  <th>Abwesenheiten</th>
  <th>Urlaub</th>
</tr>
</thead>
<tbody>
{detail_rows}
</tbody>
</table>
</body>
</html>"""
        return _Response(
            content=html,
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="statistiken_{year}.html"'},
        )


@router.get(
    "/api/export/employees",
    tags=["Export"],
    summary="Export employee list",
    description=(
        "Export all active employees as CSV or HTML.\n\n"
        "Columns: ID, Name, Vorname, Kürzel, Personalnummer, Std/Tag, Std/Woche, Std/Monat, Arbeitstage.\n\n"
        "**Required role:** Planer"
    ),
    responses={
        200: {"description": "File download (CSV/HTML)"},
        401: {"description": "Not authenticated"},
        429: {"description": "Rate limit exceeded"},
    },
)
@limiter.limit("10/minute")
def export_employees(
    request: Request,
    format: str = Query("csv"),
    _cur_user: dict = Depends(require_planer),
):
    db = get_db()
    employees = db.get_employees(include_hidden=False)
    rows = []
    for emp in employees:
        rows.append({
            "ID": emp.get('ID', ''),
            "Name": emp.get('NAME', ''),
            "Vorname": emp.get('FIRSTNAME', ''),
            "Kürzel": emp.get('SHORTNAME', ''),
            "Personalnummer": emp.get('NUMBER', ''),
            "Std/Tag": emp.get('HRSDAY', 0),
            "Std/Woche": emp.get('HRSWEEK', 0),
            "Std/Monat": emp.get('HRSMONTH', 0),
            "Arbeitstage": emp.get('WORKDAYS', ''),
        })
    if format == "html":
        headers_html = "".join(f"<th>{_html.escape(str(h))}</th>" for h in rows[0].keys()) if rows else ""
        rows_html = ""
        for i, row in enumerate(rows):
            bg = "#f8fafc" if i % 2 == 0 else "#ffffff"
            rows_html += f'<tr style="background:{bg}">' + "".join(f"<td>{_html.escape(str(v))}</td>" for v in row.values()) + "</tr>\n"
        html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<title>Mitarbeiterliste</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; }}
  h1 {{ font-size: 16px; margin-bottom: 12px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th {{ background: #1e293b; color: #fff; padding: 6px 8px; text-align: left; font-size: 10px; text-transform: uppercase; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #e2e8f0; }}
  @media print {{ @page {{ size: A4 landscape; margin: 10mm; }} }}
</style>
</head>
<body>
<h1>Mitarbeiterliste</h1>
<table><thead><tr>{headers_html}</tr></thead><tbody>
{rows_html}
</tbody></table>
</body></html>"""
        return _Response(
            content=html,
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="mitarbeiter.html"'},
        )
    if format == "xlsx":
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl nicht installiert.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mitarbeiterliste"
        thin = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        if rows:
            headers = list(rows[0].keys())
            col_widths = [20, 20, 20, 10, 18, 10, 12, 12, 14]
            for c, (h, w) in enumerate(zip(headers, col_widths), start=1):
                cell = ws.cell(1, c, h)
                cell.font = Font(bold=True, color="FFFFFF", size=9)
                cell.fill = PatternFill(fill_type="solid", fgColor="1E293B")
                cell.alignment = Alignment(horizontal="left")
                cell.border = border
                from openpyxl.utils import get_column_letter
                ws.column_dimensions[get_column_letter(c)].width = w
            for r_idx, row in enumerate(rows, start=2):
                fill_color = "F8FAFC" if r_idx % 2 == 0 else "FFFFFF"
                for c, val in enumerate(row.values(), start=1):
                    cell = ws.cell(r_idx, c, val)
                    cell.font = Font(size=9)
                    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                    cell.border = border
        buf = io.BytesIO()
        wb.save(buf)
        return _xlsx_response(buf.getvalue(), "mitarbeiter.xlsx")
    return _csv_response(rows, "mitarbeiter.csv")


@router.get(
    "/api/export/absences",
    tags=["Export"],
    summary="Export absences",
    description=(
        "Export absence entries for a given year as CSV or HTML.\n\n"
        "Optionally filter by `group_id`. Columns include employee name/short, leave type, date, and notes.\n\n"
        "**Required role:** Planer"
    ),
    responses={
        200: {"description": "File download (CSV/HTML)"},
        401: {"description": "Not authenticated"},
        429: {"description": "Rate limit exceeded"},
    },
)
@limiter.limit("10/minute")
def export_absences(
    request: Request,
    year: int = Query(...),
    group_id: Optional[int] = Query(None),
    format: str = Query("csv"),
    _cur_user: dict = Depends(require_planer),
):
    db = get_db()
    employees = db.get_employees(include_hidden=False)
    emp_map = {e['ID']: e for e in employees}
    lt_map = {lt['ID']: lt for lt in db.get_leave_types(include_hidden=True)}

    if group_id is not None:
        member_ids = set(db.get_group_members(group_id))
        emp_map = {k: v for k, v in emp_map.items() if k in member_ids}

    year_str = str(year)
    raw_absences = db._read('ABSEN')

    rows = []
    for r in raw_absences:
        d = r.get('DATE', '')
        if not (d and d.startswith(year_str)):
            continue
        eid = r.get('EMPLOYEEID')
        if eid not in emp_map:
            continue
        emp = emp_map[eid]
        ltid = r.get('LEAVETYPID')
        lt = lt_map.get(ltid) if ltid else None
        rows.append({
            "Datum": d,
            "Mitarbeiter": f"{emp.get('NAME', '')}, {emp.get('FIRSTNAME', '')}".strip(', '),
            "Kürzel": emp.get('SHORTNAME', ''),
            "Abwesenheitsart": lt.get('NAME', '') if lt else '',
            "Kürzel Art": lt.get('SHORTNAME', '') if lt else '',
        })

    rows.sort(key=lambda x: (x['Datum'], x['Mitarbeiter']))
    if format == "html":
        headers_html = "".join(f"<th>{_html.escape(str(h))}</th>" for h in rows[0].keys()) if rows else ""
        rows_html = ""
        for i, row in enumerate(rows):
            bg = "#f8fafc" if i % 2 == 0 else "#ffffff"
            rows_html += f'<tr style="background:{bg}">' + "".join(f"<td>{_html.escape(str(v))}</td>" for v in row.values()) + "</tr>\n"
        html = f"""<!DOCTYPE html>
<html lang="de">
<head>
<meta charset="utf-8">
<title>Abwesenheiten {year}</title>
<style>
  body {{ font-family: Arial, sans-serif; font-size: 11px; margin: 20px; }}
  h1 {{ font-size: 16px; margin-bottom: 12px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th {{ background: #1e293b; color: #fff; padding: 6px 8px; text-align: left; font-size: 10px; text-transform: uppercase; }}
  td {{ padding: 5px 8px; border-bottom: 1px solid #e2e8f0; }}
  @media print {{ @page {{ size: A4 portrait; margin: 10mm; }} }}
</style>
</head>
<body>
<h1>Abwesenheiten {year}</h1>
<table><thead><tr>{headers_html}</tr></thead><tbody>
{rows_html}
</tbody></table>
</body></html>"""
        return _Response(
            content=html,
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="abwesenheiten_{year}.html"'},
        )
    return _csv_response(rows, f"abwesenheiten_{year}.csv")


# ── Monatsabschluss-Report ───────────────────────────────────

@router.get(
    "/api/reports/monthly",
    tags=["Export"],
    summary="Monthly closing report (Monatsabschluss)",
    description=(
        "Generate a monthly closing report (Monatsabschluss) for all employees.\n\n"
        "- **CSV**: target/actual hours, overtime, extra-charge hours, vacation/sick days per employee\n"
        "- **PDF**: professional A4 report with table and totals row\n\n"
        "Optionally filter by `group_id`.\n\n"
        "**Required role:** Leser"
    ),
    responses={
        200: {"description": "File download (CSV/PDF)"},
        400: {"description": "Invalid month or format"},
        401: {"description": "Not authenticated"},
    },
)
def get_monthly_report(
    year: int = Query(..., description="Year (YYYY)"),
    month: int = Query(..., description="Month (1-12)"),
    format: str = Query("csv", description="Output format: csv or pdf"),
    group_id: Optional[int] = Query(None, description="Filter by group ID"),
):
    """Generate a monthly closing report (Monatsabschluss) for all employees.

    CSV: All employees with target/actual hours, overtime, extra-charge hours, vacation/sick days.
    PDF: Professional A4 report with logo placeholder, table and totals row.
    """
    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="month must be 1-12")
    if format not in ("csv", "pdf"):
        raise HTTPException(status_code=400, detail="format must be 'csv' or 'pdf'")

    from datetime import datetime as _dt

    MONTHS_DE = ["", "Januar", "Februar", "März", "April", "Mai", "Juni",
                 "Juli", "August", "September", "Oktober", "November", "Dezember"]
    month_label = f"{MONTHS_DE[month]} {year}"

    db = get_db()

    # ── Core statistics per employee ──────────────────────────
    stats = db.get_statistics(year, month, group_id=group_id)

    if not stats:
        raise HTTPException(
            status_code=404,
            detail=f"Keine Daten für {month_label} gefunden."
        )

    # ── Extra-charge hours per employee ───────────────────────
    # calculate_extracharge_hours doesn't support per-employee breakdown,
    # so we compute it per employee using the employee_id filter.
    # For performance, only compute if there are extra charges defined.
    try:
        charges_list = db.get_extracharges(include_hidden=False)
        all_charge_names: list = [c["NAME"] for c in charges_list] if charges_list else []
        xc_by_emp: dict = {}
        if charges_list:
            for s in stats:
                eid = s["employee_id"]
                emp_xc = db.calculate_extracharge_hours(year, month, employee_id=eid)
                xc_by_emp[eid] = {xc["charge_name"]: xc["hours"] for xc in emp_xc}
    except Exception:
        xc_by_emp = {}
        all_charge_names = []

    # ── Build row data ────────────────────────────────────────
    rows = []
    for s in stats:
        eid = s["employee_id"]
        row: dict = {
            "Mitarbeiter": s["employee_name"],
            "Kürzel": s["employee_short"],
            "Gruppe": s.get("group_name", ""),
            "Soll-Std.": s["target_hours"],
            "Ist-Std.": s["actual_hours"],
            "Überstunden": s["overtime_hours"],
            "Dienste": s["shifts_count"],
            "Abwesenheitstage": s["absence_days"],
            "Urlaubstage": s["vacation_used"],
            "Kranktage": s.get("sick_days", 0),
        }
        for cn in all_charge_names:
            row[f"Zuschlag: {cn}"] = round(xc_by_emp.get(eid, {}).get(cn, 0.0), 2)
        rows.append(row)

    filename_base = f"monatsabschluss_{year}_{month:02d}"

    # ── CSV output ────────────────────────────────────────────
    if format == "csv":
        buf = io.StringIO()
        if rows:
            writer = csv.DictWriter(buf, fieldnames=rows[0].keys(), lineterminator='\r\n')
            writer.writeheader()
            writer.writerows(rows)
        content = buf.getvalue()
        return _Response(
            content=content.encode("utf-8-sig"),  # BOM for Excel compatibility
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    # ── PDF output ────────────────────────────────────────────
    try:
        from fpdf import FPDF
    except ImportError:
        raise HTTPException(status_code=500, detail="fpdf2 nicht installiert. Bitte 'pip install fpdf2' ausführen.")

    class SP5Report(FPDF):
        def header(self):
            # Logo placeholder
            self.set_fill_color(30, 41, 59)  # dark slate
            self.rect(10, 8, 30, 16, 'F')
            self.set_font("Helvetica", "B", 9)
            self.set_text_color(255, 255, 255)
            self.set_xy(10, 11)
            self.cell(30, 6, "SP5", align="C")
            self.set_xy(10, 17)
            self.cell(30, 4, "LOGO", align="C")

            # Title
            self.set_text_color(30, 41, 59)
            self.set_font("Helvetica", "B", 13)
            self.set_xy(44, 9)
            self.cell(0, 7, "Monatsabschluss-Report", ln=0)
            self.set_font("Helvetica", "", 9)
            self.set_xy(44, 16)
            self.cell(0, 5, f"Zeitraum: {month_label}  |  Erstellt: {_dt.now().strftime('%d.%m.%Y %H:%M')}", ln=0)
            self.set_draw_color(30, 41, 59)
            self.set_line_width(0.5)
            self.line(10, 26, self.w - 10, 26)
            self.ln(20)

        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "I", 7)
            self.set_text_color(120, 120, 120)
            self.cell(0, 5, f"OpenSchichtplaner5  |  {month_label}  |  Seite {self.page_no()}", align="C")

    pdf = SP5Report(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    # ── Summary box ───────────────────────────────────────────
    total_soll = sum(s["target_hours"] for s in stats)
    total_ist = sum(s["actual_hours"] for s in stats)
    total_ot = round(total_ist - total_soll, 2)
    total_abs = sum(s["absence_days"] for s in stats)
    total_vac = sum(s["vacation_used"] for s in stats)
    total_sick = sum(s.get("sick_days", 0) for s in stats)

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(241, 245, 249)
    pdf.set_draw_color(203, 213, 225)
    pdf.set_line_width(0.3)

    box_y = pdf.get_y()
    kpi_labels = [
        ("Mitarbeiter", str(len(stats))),
        ("Soll-Std.", f"{total_soll:.1f} h"),
        ("Ist-Std.", f"{total_ist:.1f} h"),
        ("Überstunden", f"{'+' if total_ot >= 0 else ''}{total_ot:.1f} h"),
        ("Urlaubstage", str(total_vac)),
        ("Kranktage", str(total_sick)),
        ("Abwesenheiten", str(total_abs)),
    ]
    box_w = (pdf.w - 20) / len(kpi_labels)
    for i, (lbl, val) in enumerate(kpi_labels):
        x = 10 + i * box_w
        pdf.set_xy(x, box_y)
        pdf.set_fill_color(241, 245, 249)
        pdf.rect(x, box_y, box_w - 1, 14, 'FD')
        pdf.set_font("Helvetica", "", 7)
        pdf.set_text_color(100, 116, 139)
        pdf.set_xy(x, box_y + 1)
        pdf.cell(box_w - 1, 5, lbl, align="C")
        pdf.set_font("Helvetica", "B", 10)
        pdf.set_text_color(30, 41, 59)
        if lbl == "Überstunden":
            ot_val = total_ot
            pdf.set_text_color(22, 163, 74 if ot_val >= 0 else 220, )
            if ot_val < 0:
                pdf.set_text_color(220, 38, 38)
        pdf.set_xy(x, box_y + 6)
        pdf.cell(box_w - 1, 7, val, align="C")

    pdf.set_y(box_y + 17)
    pdf.ln(2)

    # ── Table header ──────────────────────────────────────────
    # Fixed columns + dynamic surcharge columns
    fixed_cols = [
        ("Mitarbeiter", 42),
        ("Kürzel", 14),
        ("Gruppe", 28),
        ("Soll h", 16),
        ("Ist h", 16),
        ("ÜSt h", 16),
        ("Dienste", 14),
        ("Abw.", 12),
        ("Url.", 12),
        ("Krank", 12),
    ]
    # Dynamic surcharge cols (max 3 displayed to fit page)
    xc_cols = [(f"ZZ: {cn[:8]}", 16) for cn in all_charge_names[:3]]
    all_cols = fixed_cols + xc_cols

    # Scale if too wide
    total_w = sum(w for _, w in all_cols)
    avail_w = pdf.w - 20
    scale = avail_w / total_w if total_w > avail_w else 1.0
    all_cols = [(label, round(w * scale, 1)) for label, w in all_cols]

    ROW_H = 7
    HDR_H = 9

    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 7)
    pdf.set_draw_color(100, 116, 139)
    pdf.set_line_width(0.2)

    pdf.get_y()
    for label, w in all_cols:
        pdf.cell(w, HDR_H, label, border=1, fill=True, align="C")
    pdf.ln()

    # ── Table rows ────────────────────────────────────────────
    pdf.set_font("Helvetica", "", 7)
    for i, row in enumerate(rows):
        eid = stats[i]["employee_id"]
        if pdf.get_y() + ROW_H > pdf.h - 22:
            pdf.add_page()

        fill = i % 2 == 0
        if fill:
            pdf.set_fill_color(248, 250, 252)
        else:
            pdf.set_fill_color(255, 255, 255)
        pdf.set_text_color(30, 41, 59)

        col_idx = 0
        # Mitarbeiter
        pdf.cell(all_cols[col_idx][1], ROW_H, str(row["Mitarbeiter"])[:24], border=1, fill=True, align="L"); col_idx += 1  # noqa: E702
        # Kürzel
        pdf.cell(all_cols[col_idx][1], ROW_H, str(row["Kürzel"]), border=1, fill=True, align="C"); col_idx += 1  # noqa: E702
        # Gruppe
        pdf.cell(all_cols[col_idx][1], ROW_H, str(row["Gruppe"])[:16], border=1, fill=True, align="L"); col_idx += 1  # noqa: E702
        # Soll h
        pdf.cell(all_cols[col_idx][1], ROW_H, f"{row['Soll-Std.']:.1f}", border=1, fill=True, align="R"); col_idx += 1  # noqa: E702
        # Ist h
        pdf.cell(all_cols[col_idx][1], ROW_H, f"{row['Ist-Std.']:.1f}", border=1, fill=True, align="R"); col_idx += 1  # noqa: E702
        # ÜSt h — color
        ot_val = row["Überstunden"]
        if ot_val > 0:
            pdf.set_text_color(22, 163, 74)
        elif ot_val < 0:
            pdf.set_text_color(220, 38, 38)
        else:
            pdf.set_text_color(100, 116, 139)
        ot_sign = "+" if ot_val > 0 else ""
        pdf.cell(all_cols[col_idx][1], ROW_H, f"{ot_sign}{ot_val:.1f}", border=1, fill=True, align="R")
        pdf.set_text_color(30, 41, 59)
        col_idx += 1
        # Dienste
        pdf.cell(all_cols[col_idx][1], ROW_H, str(row["Dienste"]), border=1, fill=True, align="C"); col_idx += 1  # noqa: E702
        # Abw.
        pdf.cell(all_cols[col_idx][1], ROW_H, str(row["Abwesenheitstage"]), border=1, fill=True, align="C"); col_idx += 1  # noqa: E702
        # Url.
        pdf.cell(all_cols[col_idx][1], ROW_H, str(row["Urlaubstage"]), border=1, fill=True, align="C"); col_idx += 1  # noqa: E702
        # Krank
        pdf.cell(all_cols[col_idx][1], ROW_H, str(row["Kranktage"]), border=1, fill=True, align="C"); col_idx += 1  # noqa: E702
        # Surcharge columns
        for cn in all_charge_names[:3]:
            hrs = row.get(f"Zuschlag: {cn}", 0.0)
            pdf.cell(all_cols[col_idx][1], ROW_H, f"{hrs:.1f}" if hrs else "-", border=1, fill=True, align="C")
            col_idx += 1
        pdf.ln()

    # ── Totals row ────────────────────────────────────────────
    if pdf.get_y() + ROW_H + 2 > pdf.h - 22:
        pdf.add_page()
    pdf.set_fill_color(30, 41, 59)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 7)
    totals_row = {
        "Mitarbeiter": f"Gesamt ({len(stats)} Mitarbeiter)",
        "Kürzel": "",
        "Gruppe": "",
        "Soll-Std.": f"{total_soll:.1f}",
        "Ist-Std.": f"{total_ist:.1f}",
        "Überstunden": f"{'+' if total_ot >= 0 else ''}{total_ot:.1f}",
        "Dienste": str(sum(s["shifts_count"] for s in stats)),
        "Abwesenheitstage": str(total_abs),
        "Urlaubstage": str(total_vac),
        "Kranktage": str(total_sick),
    }
    col_idx = 0
    pdf.cell(all_cols[col_idx][1], ROW_H, totals_row["Mitarbeiter"][:30], border=1, fill=True, align="L"); col_idx += 1  # noqa: E702
    pdf.cell(all_cols[col_idx][1], ROW_H, "", border=1, fill=True); col_idx += 1  # noqa: E702
    pdf.cell(all_cols[col_idx][1], ROW_H, "", border=1, fill=True); col_idx += 1  # noqa: E702
    pdf.cell(all_cols[col_idx][1], ROW_H, totals_row["Soll-Std."], border=1, fill=True, align="R"); col_idx += 1  # noqa: E702
    pdf.cell(all_cols[col_idx][1], ROW_H, totals_row["Ist-Std."], border=1, fill=True, align="R"); col_idx += 1  # noqa: E702
    pdf.cell(all_cols[col_idx][1], ROW_H, totals_row["Überstunden"], border=1, fill=True, align="R"); col_idx += 1  # noqa: E702
    pdf.cell(all_cols[col_idx][1], ROW_H, totals_row["Dienste"], border=1, fill=True, align="C"); col_idx += 1  # noqa: E702
    pdf.cell(all_cols[col_idx][1], ROW_H, totals_row["Abwesenheitstage"], border=1, fill=True, align="C"); col_idx += 1  # noqa: E702
    pdf.cell(all_cols[col_idx][1], ROW_H, totals_row["Urlaubstage"], border=1, fill=True, align="C"); col_idx += 1  # noqa: E702
    pdf.cell(all_cols[col_idx][1], ROW_H, totals_row["Kranktage"], border=1, fill=True, align="C"); col_idx += 1  # noqa: E702
    for cn in all_charge_names[:3]:
        total_xc = round(sum(xc_by_emp.get(s["employee_id"], {}).get(cn, 0.0) for s in stats), 1)
        pdf.cell(all_cols[col_idx][1], ROW_H, f"{total_xc:.1f}" if total_xc else "-", border=1, fill=True, align="C")
        col_idx += 1
    pdf.ln()

    pdf_bytes = pdf.output()
    return _Response(
        content=bytes(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'},
    )


# ── Zeitkonto / Überstunden ──────────────────────────────────

@router.get(
    "/api/zeitkonto",
    tags=["Statistics"],
    summary="Time balance (Zeitkonto)",
    description=(
        "Return yearly time balance for all employees (or filtered by group/employee).\n\n"
        "Each row contains target hours, actual hours, and the running saldo per month.\n\n"
        "**Required role:** Leser"
    ),
)
def get_zeitkonto(
    year: int = Query(..., description="Year"),
    group_id: Optional[int] = Query(None, description="Filter by group ID"),
    employee_id: Optional[int] = Query(None, description="Filter by employee ID"),
):
    return get_db().get_zeitkonto(year=year, group_id=group_id, employee_id=employee_id)


@router.get(
    "/api/zeitkonto/detail",
    tags=["Statistics"],
    summary="Time balance detail for one employee",
    description=(
        "Return the full monthly time-balance calculation for a single employee in a given year.\n\n"
        "Returns 404 if the employee is not found.\n\n"
        "**Required role:** Leser"
    ),
    responses={
        404: {"description": "Employee not found"},
    },
)
def get_zeitkonto_detail(
    year: int = Query(..., description="Year"),
    employee_id: int = Query(..., description="Employee ID"),
):
    db = get_db()
    result = db.calculate_time_balance(employee_id=employee_id, year=year)
    if not result:
        raise HTTPException(status_code=404, detail="Employee not found")
    return result


@router.get(
    "/api/zeitkonto/summary",
    tags=["Statistics"],
    summary="Time balance summary (team totals)",
    description=(
        "Return aggregated time-balance totals for a year (optionally filtered by group).\n\n"
        "Includes total target/actual hours, overall saldo, and counts of employees with positive/negative saldo.\n\n"
        "**Required role:** Leser"
    ),
)
def get_zeitkonto_summary(
    year: int = Query(..., description="Year"),
    group_id: Optional[int] = Query(None, description="Filter by group ID"),
):
    rows = get_db().get_zeitkonto(year=year, group_id=group_id)
    total_target = sum(r['total_target_hours'] for r in rows)
    total_actual = sum(r['total_actual_hours'] for r in rows)
    total_saldo = sum(r['total_saldo'] for r in rows)
    pos = sum(1 for r in rows if r['total_saldo'] >= 0)
    neg = len(rows) - pos
    return {
        'year': year,
        'group_id': group_id,
        'employee_count': len(rows),
        'total_target_hours': round(total_target, 2),
        'total_actual_hours': round(total_actual, 2),
        'total_saldo': round(total_saldo, 2),
        'positive_count': pos,
        'negative_count': neg,
    }


@router.get(
    "/api/bookings",
    tags=["Statistics"],
    summary="List manual hour bookings",
    description=(
        "Return manual hour bookings (Stundenbuchungen). "
        "Filter by `year`, `month`, and/or `employee_id`.\n\n"
        "**Required role:** Leser"
    ),
)
def get_bookings(
    year: Optional[int] = Query(None, description="Filter by year"),
    month: Optional[int] = Query(None, description="Filter by month (1-12), use with year"),
    employee_id: Optional[int] = Query(None, description="Filter by employee ID"),
):
    return get_db().get_bookings(year=year, month=month, employee_id=employee_id)


class BookingCreate(BaseModel):
    employee_id: int = Field(..., gt=0)
    date: str = Field(..., pattern=r'^\d{4}-\d{2}-\d{2}$')
    type: int = Field(0, ge=0, le=1)   # 0 = Iststundenkonto, 1 = Sollstundenkonto
    value: float
    note: Optional[str] = Field('', max_length=500)


@router.post(
    "/api/bookings",
    tags=["Statistics"],
    summary="Create manual hour booking",
    description=(
        "Create a manual hour booking (Stundenbuchung) for an employee.\n\n"
        "`type`: 0 = Ist-Stunden, 1 = Soll-Stunden\n\n"
        "**Required role:** Planer"
    ),
    responses={
        400: {"description": "Invalid date format or type"},
        401: {"description": "Not authenticated"},
    },
)
def create_booking(body: BookingCreate, _cur_user: dict = Depends(require_planer)):
    try:
        from datetime import datetime
        datetime.strptime(body.date, '%Y-%m-%d')
    except ValueError:
        raise HTTPException(status_code=400, detail="Ungültiges Datumsformat, bitte JJJJ-MM-TT verwenden")
    if body.type not in (0, 1):
        raise HTTPException(status_code=400, detail="type must be 0 (Ist) or 1 (Soll)")
    try:
        result = get_db().create_booking(
            employee_id=body.employee_id,
            date_str=body.date,
            booking_type=body.type,
            value=body.value,
            note=body.note or '',
        )
        return {"ok": True, "record": result}
    except Exception as e:
        raise _sanitize_500(e)


@router.delete(
    "/api/bookings/{booking_id}",
    tags=["Statistics"],
    summary="Delete manual hour booking",
    description="Delete a manual hour booking by ID. Returns 404 if not found.\n\n**Required role:** Planer",
    responses={
        404: {"description": "Booking not found"},
        401: {"description": "Not authenticated"},
    },
)
def delete_booking(booking_id: int, _cur_user: dict = Depends(require_planer)):
    try:
        count = get_db().delete_booking(booking_id)
        if count == 0:
            raise HTTPException(status_code=404, detail="Booking not found")
        return {"ok": True, "deleted": booking_id}
    except HTTPException:
        raise
    except Exception as e:
        raise _sanitize_500(e)


# ── Carry Forward (Saldo-Übertrag) ────────────────────────────

@router.get(
    "/api/bookings/carry-forward",
    tags=["Statistics"],
    summary="Get carry-forward saldo",
)
def get_carry_forward(employee_id: int = Query(...), year: int = Query(...)):
    try:
        return get_db().get_carry_forward(employee_id=employee_id, year=year)
    except Exception as e:
        raise _sanitize_500(e)


class CarryForwardSet(BaseModel):
    employee_id: int = Field(..., gt=0)
    year: int = Field(..., ge=2000, le=2100)
    hours: float


@router.post(
    "/api/bookings/carry-forward",
    tags=["Statistics"],
    summary="Set carry-forward saldo",
)
def set_carry_forward(body: CarryForwardSet, _cur_user: dict = Depends(require_planer)):
    try:
        result = get_db().set_carry_forward(
            employee_id=body.employee_id,
            year=body.year,
            hours=body.hours,
        )
        return {"ok": True, "record": result}
    except Exception as e:
        raise _sanitize_500(e)


class AnnualStatementBody(BaseModel):
    employee_id: int = Field(..., gt=0)
    year: int = Field(..., ge=2000, le=2100)


@router.post(
    "/api/bookings/annual-statement",
    tags=["Statistics"],
    summary="Generate annual statement",
)
def annual_statement(body: AnnualStatementBody, _cur_user: dict = Depends(require_planer)):
    try:
        result = get_db().calculate_annual_statement(
            employee_id=body.employee_id,
            year=body.year,
        )
        return {"ok": True, "result": result}
    except Exception as e:
        raise _sanitize_500(e)


@router.get(
    "/api/overtime-records",
    tags=["Statistics"],
    summary="List overtime records",
)
def get_overtime_records(
    year: Optional[int] = Query(None, description="Filter by year"),
    employee_id: Optional[int] = Query(None, description="Filter by employee ID"),
):
    return get_db().get_overtime_records(year=year, employee_id=employee_id)


def _decode_csv(content: bytes) -> str:
    """Try UTF-8 with BOM first, then latin-1."""
    try:
        return content.decode('utf-8-sig')
    except UnicodeDecodeError:
        return content.decode('latin-1')


@router.post(
    "/api/import/employees",
    tags=["Import"],
    summary="Import employees from CSV",
)
async def import_employees(file: UploadFile = File(...)):
    """Import employees from CSV. Required columns: NAME or NACHNAME.
    Accepted column aliases: VORNAME/FIRSTNAME, NACHNAME/NAME, KURZZEICHEN/SHORTNAME,
    NUMBER/PERSONALNUMMER, HRSDAY, HRSWEEK, HRSMONTH, SEX."""
    content = await file.read()
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    errors = []
    db = get_db()

    for i, row in enumerate(reader, start=2):  # row 1 = header
        # Normalize keys
        row = {k.strip().upper(): v.strip() for k, v in row.items() if k}

        # Alias mapping
        name = row.get('NAME') or row.get('NACHNAME') or ''
        firstname = row.get('FIRSTNAME') or row.get('VORNAME') or ''
        shortname = row.get('SHORTNAME') or row.get('KURZZEICHEN') or ''
        number = row.get('NUMBER') or row.get('PERSONALNUMMER') or ''

        if not name:
            errors.append(f"Zeile {i}: NAME/NACHNAME fehlt — übersprungen")
            skipped += 1
            continue

        try:
            data = {
                'NAME': name,
                'FIRSTNAME': firstname,
                'SHORTNAME': shortname,
                'NUMBER': number,
                'SEX': int(row.get('SEX') or 0),
                'HRSDAY': float(row.get('HRSDAY') or 0),
                'HRSWEEK': float(row.get('HRSWEEK') or 0),
                'HRSMONTH': float(row.get('HRSMONTH') or 0),
                'WORKDAYS': row.get('WORKDAYS') or '1 1 1 1 1 0 0 0',
                'HIDE': False,
            }
            db.create_employee(data)
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i} ({name}): {e}")

    return {"imported": imported, "errors": errors, "skipped": skipped}


@router.post(
    "/api/import/shifts",
    tags=["Import"],
    summary="Import shifts from CSV",
)
async def import_shifts(file: UploadFile = File(...)):
    """Import shifts from CSV. Required: NAME.
    Optional: KURZZEICHEN/SHORTNAME, FARBE/COLORBK (hex #RRGGBB or int BGR), DURATION0."""
    content = await file.read()
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    errors = []
    db = get_db()

    def _parse_color(val: str) -> int:
        """Parse #RRGGBB hex to BGR int, or pass-through int."""
        if not val:
            return 16777215  # white
        val = val.strip()
        if val.startswith('#') and len(val) == 7:
            try:
                r = int(val[1:3], 16)
                g = int(val[3:5], 16)
                b = int(val[5:7], 16)
                return (b << 16) | (g << 8) | r
            except ValueError:
                return 16777215
        try:
            return int(val)
        except ValueError:
            return 16777215

    for i, row in enumerate(reader, start=2):
        row = {k.strip().upper(): v.strip() for k, v in row.items() if k}

        name = row.get('NAME') or ''
        if not name:
            errors.append(f"Zeile {i}: NAME fehlt — übersprungen")
            skipped += 1
            continue

        shortname = row.get('SHORTNAME') or row.get('KURZZEICHEN') or ''
        colorbk_raw = row.get('COLORBK') or row.get('FARBE') or row.get('HINTERGRUNDFARBE') or ''
        colortext_raw = row.get('COLORTEXT') or row.get('TEXTFARBE') or ''

        try:
            data = {
                'NAME': name,
                'SHORTNAME': shortname,
                'COLORBK': _parse_color(colorbk_raw),
                'COLORTEXT': _parse_color(colortext_raw) if colortext_raw else 0,
                'COLORBAR': 0,
                'DURATION0': float(row.get('DURATION0') or row.get('DAUER') or 0),
                'HIDE': False,
            }
            db.create_shift(data)
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i} ({name}): {e}")

    return {"imported": imported, "errors": errors, "skipped": skipped}


@router.post(
    "/api/import/absences",
    tags=["Import"],
    summary="Import absences from CSV",
)
async def import_absences(file: UploadFile = File(...)):
    """Import absences from CSV. Required: EMPLOYEE_ID, DATE (YYYY-MM-DD), LEAVE_TYPE_ID."""
    content = await file.read()
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    errors = []
    db = get_db()

    for i, row in enumerate(reader, start=2):
        row = {k.strip().upper(): v.strip() for k, v in row.items() if k}

        emp_id_raw = row.get('EMPLOYEE_ID') or row.get('MITARBEITER_ID') or ''
        date_raw = row.get('DATE') or row.get('DATUM') or ''
        lt_id_raw = row.get('LEAVE_TYPE_ID') or row.get('ABWESENHEITSART_ID') or ''

        if not emp_id_raw or not date_raw or not lt_id_raw:
            errors.append(f"Zeile {i}: Pflichtfelder fehlen (EMPLOYEE_ID, DATE, LEAVE_TYPE_ID) — übersprungen")
            skipped += 1
            continue

        try:
            from datetime import datetime
            datetime.strptime(date_raw, '%Y-%m-%d')
        except ValueError:
            errors.append(f"Zeile {i}: Ungültiges Datum '{date_raw}' (erwartet YYYY-MM-DD) — übersprungen")
            skipped += 1
            continue

        try:
            emp_id = int(emp_id_raw)
            lt_id = int(lt_id_raw)
            db.add_absence(emp_id, date_raw, lt_id)
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i}: {e}")

    return {"imported": imported, "errors": errors, "skipped": skipped}


@router.post(
    "/api/import/holidays",
    tags=["Import"],
    summary="Import holidays from CSV",
)
async def import_holidays(file: UploadFile = File(...)):
    """Import holidays from CSV. Required: DATE (YYYY-MM-DD), NAME.
    Optional: INTERVAL (0=einmalig, 1=jährlich), REGION (ignored, for info only)."""
    content = await file.read()
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    errors = []
    db = get_db()

    for i, row in enumerate(reader, start=2):
        row = {k.strip().upper(): v.strip() for k, v in row.items() if k}

        date_raw = row.get('DATE') or row.get('DATUM') or ''
        name = row.get('NAME') or row.get('BEZEICHNUNG') or ''

        if not date_raw or not name:
            errors.append(f"Zeile {i}: DATE und NAME sind Pflicht — übersprungen")
            skipped += 1
            continue

        try:
            from datetime import datetime
            datetime.strptime(date_raw, '%Y-%m-%d')
        except ValueError:
            errors.append(f"Zeile {i}: Ungültiges Datum '{date_raw}' (erwartet YYYY-MM-DD) — übersprungen")
            skipped += 1
            continue

        try:
            interval_raw = row.get('INTERVAL') or row.get('JAEHRLICH') or '0'
            data = {
                'DATE': date_raw,
                'NAME': name,
                'INTERVAL': int(interval_raw) if interval_raw.isdigit() else 0,
            }
            db.create_holiday(data)
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i} ({name}): {e}")

    return {"imported": imported, "errors": errors, "skipped": skipped}


@router.post(
    "/api/import/bookings-actual",
    tags=["Import"],
    summary="Import actual-hour bookings from CSV",
)
async def import_bookings_actual(file: UploadFile = File(...)):
    """Import actual-hour bookings (TYPE=0) from CSV.
    Required: Personalnummer,Datum,Stunden. Optional: Notiz."""
    content = await file.read()
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    errors = []
    db = get_db()

    emp_by_number = {(str(e.get('NUMBER', '')) or '').strip(): e for e in db.get_employees(include_hidden=False)}

    for i, row in enumerate(reader, start=2):
        row = {k.strip().upper(): v.strip() for k, v in row.items() if k}
        nummer = row.get('PERSONALNUMMER') or row.get('NUMBER') or ''
        date_raw = row.get('DATUM') or row.get('DATE') or ''
        stunden_raw = row.get('STUNDEN') or row.get('HOURS') or ''
        notiz = row.get('NOTIZ') or row.get('NOTE') or ''

        if not nummer or not date_raw or not stunden_raw:
            errors.append(f"Zeile {i}: Pflichtfelder fehlen (Personalnummer,Datum,Stunden) — übersprungen")
            skipped += 1
            continue

        emp = emp_by_number.get(nummer)
        if not emp:
            errors.append(f"Zeile {i}: Personalnummer '{nummer}' nicht gefunden — übersprungen")
            skipped += 1
            continue

        try:
            from datetime import datetime as _dt
            _dt.strptime(date_raw, '%Y-%m-%d')
            stunden = float(stunden_raw.replace(',', '.'))
            db.create_booking(emp['ID'], date_raw, 0, stunden, notiz)
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i}: {e}")
            skipped += 1

    return {"imported": imported, "skipped": skipped, "errors": errors}


@router.post(
    "/api/import/bookings-nominal",
    tags=["Import"],
    summary="Import nominal-hour bookings from CSV",
)
async def import_bookings_nominal(file: UploadFile = File(...)):
    """Import nominal-hour bookings (TYPE=1) from CSV.
    Required: Personalnummer,Datum,Stunden. Optional: Notiz."""
    content = await file.read()
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    errors = []
    db = get_db()

    emp_by_number = {(str(e.get('NUMBER', '')) or '').strip(): e for e in db.get_employees(include_hidden=False)}

    for i, row in enumerate(reader, start=2):
        row = {k.strip().upper(): v.strip() for k, v in row.items() if k}
        nummer = row.get('PERSONALNUMMER') or row.get('NUMBER') or ''
        date_raw = row.get('DATUM') or row.get('DATE') or ''
        stunden_raw = row.get('STUNDEN') or row.get('HOURS') or ''
        notiz = row.get('NOTIZ') or row.get('NOTE') or ''

        if not nummer or not date_raw or not stunden_raw:
            errors.append(f"Zeile {i}: Pflichtfelder fehlen (Personalnummer,Datum,Stunden) — übersprungen")
            skipped += 1
            continue

        emp = emp_by_number.get(nummer)
        if not emp:
            errors.append(f"Zeile {i}: Personalnummer '{nummer}' nicht gefunden — übersprungen")
            skipped += 1
            continue

        try:
            from datetime import datetime as _dt
            _dt.strptime(date_raw, '%Y-%m-%d')
            stunden = float(stunden_raw.replace(',', '.'))
            db.create_booking(emp['ID'], date_raw, 1, stunden, notiz)
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i}: {e}")
            skipped += 1

    return {"imported": imported, "skipped": skipped, "errors": errors}


@router.post(
    "/api/import/entitlements",
    tags=["Import"],
    summary="Import vacation entitlements from CSV",
)
async def import_entitlements(file: UploadFile = File(...)):
    """Import leave entitlements from CSV.
    Required: Personalnummer,Jahr,Abwesenheitsart-Kürzel,Tage."""
    content = await file.read()
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    errors = []
    db = get_db()

    emp_by_number = {(str(e.get('NUMBER', '')) or '').strip(): e for e in db.get_employees(include_hidden=False)}
    lt_by_short = {lt['SHORTNAME'].strip().upper(): lt for lt in db.get_leave_types(include_hidden=False)}

    for i, row in enumerate(reader, start=2):
        row = {k.strip().upper(): v.strip() for k, v in row.items() if k}
        nummer = row.get('PERSONALNUMMER') or row.get('NUMBER') or ''
        year_raw = row.get('JAHR') or row.get('YEAR') or ''
        kuerzel = (row.get('ABWESENHEITSART') or row.get('KÜRZEL') or row.get('KURZEL') or row.get('SHORTNAME') or '').upper()
        tage_raw = row.get('TAGE') or row.get('DAYS') or ''

        if not nummer or not year_raw or not kuerzel or not tage_raw:
            errors.append(f"Zeile {i}: Pflichtfelder fehlen (Personalnummer,Jahr,Abwesenheitsart-Kürzel,Tage) — übersprungen")
            skipped += 1
            continue

        emp = emp_by_number.get(nummer)
        if not emp:
            errors.append(f"Zeile {i}: Personalnummer '{nummer}' nicht gefunden — übersprungen")
            skipped += 1
            continue

        lt = lt_by_short.get(kuerzel)
        if not lt:
            errors.append(f"Zeile {i}: Abwesenheitsart-Kürzel '{kuerzel}' nicht gefunden — übersprungen")
            skipped += 1
            continue

        try:
            year = int(year_raw)
            tage = float(tage_raw.replace(',', '.'))
            db.set_leave_entitlement(emp['ID'], year, tage, leave_type_id=lt['ID'])
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i}: {e}")
            skipped += 1

    return {"imported": imported, "skipped": skipped, "errors": errors}


@router.post(
    "/api/import/absences-csv",
    tags=["Import"],
    summary="Import absences from CSV (alternate format)",
)
async def import_absences_csv(file: UploadFile = File(...)):
    """Import absences from CSV using Personalnummer and Abwesenheitsart-Kürzel.
    Required: Personalnummer,Datum,Abwesenheitsart-Kürzel."""
    content = await file.read()
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    errors = []
    db = get_db()

    emp_by_number = {(str(e.get('NUMBER', '')) or '').strip(): e for e in db.get_employees(include_hidden=False)}
    lt_by_short = {lt['SHORTNAME'].strip().upper(): lt for lt in db.get_leave_types(include_hidden=False)}

    for i, row in enumerate(reader, start=2):
        row = {k.strip().upper(): v.strip() for k, v in row.items() if k}
        nummer = row.get('PERSONALNUMMER') or row.get('NUMBER') or ''
        date_raw = row.get('DATUM') or row.get('DATE') or ''
        kuerzel = (row.get('ABWESENHEITSART') or row.get('KÜRZEL') or row.get('KURZEL') or row.get('SHORTNAME') or '').upper()

        if not nummer or not date_raw or not kuerzel:
            errors.append(f"Zeile {i}: Pflichtfelder fehlen (Personalnummer,Datum,Abwesenheitsart-Kürzel) — übersprungen")
            skipped += 1
            continue

        emp = emp_by_number.get(nummer)
        if not emp:
            errors.append(f"Zeile {i}: Personalnummer '{nummer}' nicht gefunden — übersprungen")
            skipped += 1
            continue

        lt = lt_by_short.get(kuerzel)
        if not lt:
            errors.append(f"Zeile {i}: Abwesenheitsart-Kürzel '{kuerzel}' nicht gefunden — übersprungen")
            skipped += 1
            continue

        try:
            from datetime import datetime as _dt
            _dt.strptime(date_raw, '%Y-%m-%d')
            db.add_absence(emp['ID'], date_raw, lt['ID'])
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i}: {e}")
            skipped += 1

    return {"imported": imported, "skipped": skipped, "errors": errors}


@router.post(
    "/api/import/groups",
    tags=["Import"],
    summary="Import groups from CSV",
)
async def import_groups(file: UploadFile = File(...)):
    """Import groups from CSV.
    Required: Name. Optional: Kürzel, Übergeordnete-Gruppe-Name."""
    content = await file.read()
    text = _decode_csv(content)
    reader = csv.DictReader(io.StringIO(text))

    imported = 0
    skipped = 0
    errors = []
    db = get_db()

    existing_groups = db.get_groups(include_hidden=True)
    group_by_name = {g['NAME'].strip().upper(): g for g in existing_groups}

    for i, row in enumerate(reader, start=2):
        row = {k.strip().upper(): v.strip() for k, v in row.items() if k}
        name = row.get('NAME') or row.get('BEZEICHNUNG') or ''
        kuerzel = row.get('KÜRZEL') or row.get('KURZEL') or row.get('SHORTNAME') or ''
        parent_name = (row.get('ÜBERGEORDNETE-GRUPPE-NAME') or row.get('UEBERGEORDNETE-GRUPPE-NAME') or
                       row.get('PARENT') or row.get('SUPERGRUPPE') or '').strip().upper()

        if not name:
            errors.append(f"Zeile {i}: NAME fehlt — übersprungen")
            skipped += 1
            continue

        parent_id = None
        if parent_name:
            parent_grp = group_by_name.get(parent_name)
            if not parent_grp:
                errors.append(f"Zeile {i}: Übergeordnete Gruppe '{parent_name}' nicht gefunden — übersprungen")
                skipped += 1
                continue
            parent_id = parent_grp['ID']

        try:
            data = {
                'NAME': name,
                'SHORTNAME': kuerzel,
                'SUPERID': parent_id or 0,
                'HIDE': False,
            }
            db.create_group(data)
            # Refresh for subsequent lookups
            group_by_name[name.upper()] = {'NAME': name, 'ID': -1}
            imported += 1
        except Exception as e:
            errors.append(f"Zeile {i} ({name}): {e}")
            skipped += 1

    return {"imported": imported, "skipped": skipped, "errors": errors}


# ── Burnout-Radar ────────────────────────────────────────────

@router.get(
    "/api/burnout-radar",
    tags=["Statistics"],
    summary="Burnout radar analysis",
)
def get_burnout_radar(
    year: int = Query(..., description="Year"),
    month: int = Query(..., description="Month 1-12"),
    streak_threshold: int = Query(6, description="Min consecutive days to flag"),
    overtime_threshold_pct: float = Query(20.0, description="Min overtime % to flag"),
    group_id: Optional[int] = Query(None, description="Filter by group"),
):
    """Return list of at-risk employees (long streaks or significant overtime)."""
    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="Ungültiger Monat: muss zwischen 1 und 12 liegen")
    if not (2000 <= year <= 2100):
        raise HTTPException(status_code=400, detail="Ungültiges Jahr: muss zwischen 2000 und 2100 liegen")
    return get_db().get_burnout_radar(
        year=year,
        month=month,
        streak_threshold=streak_threshold,
        overtime_threshold_pct=overtime_threshold_pct,
        group_id=group_id,
    )




# ── Überstunden-Zusammenfassung ───────────────────────────────

@router.get(
    "/api/overtime-summary",
    tags=["Statistics"],
    summary="Overtime summary",
)
def get_overtime_summary(
    year: Optional[int] = Query(None, description="Year (YYYY), defaults to current year"),
    group_id: Optional[int] = Query(None, description="Filter by group"),
):
    """Return overtime summary (Überstunden) per employee for a given year."""
    from datetime import date as _date
    if year is None:
        year = _date.today().year
    rows = get_db().get_overtime_summary(year=year, group_id=group_id)
    total_soll = sum(r['soll'] for r in rows)
    total_ist = sum(r['ist'] for r in rows)
    total_delta = round(total_ist - total_soll, 2)
    plus_count = sum(1 for r in rows if r['delta'] >= 0)
    minus_count = sum(1 for r in rows if r['delta'] < 0)
    return {
        'year': year,
        'group_id': group_id,
        'employees': rows,
        'summary': {
            'total_soll': round(total_soll, 2),
            'total_ist': round(total_ist, 2),
            'total_delta': total_delta,
            'plus_count': plus_count,
            'minus_count': minus_count,
            'employee_count': len(rows),
        },
    }


# ── Warnings Center ──────────────────────────────────────────

@router.get(
    "/api/warnings",
    tags=["Statistics"],
    summary="Schedule warnings and anomalies",
)
def get_warnings(
    year: Optional[int] = Query(None, description="Year (YYYY), defaults to current year"),
    month: Optional[int] = Query(None, description="Month (1-12), defaults to current month"),
    _cur_user: dict = Depends(require_auth),
):
    """Return a list of active warnings for the Warnings Center.

    Warning types:
    - next_month_unplanned: Next month not yet scheduled (< 7 days until month end)
    - overtime_exceeded: Employee has overtime > threshold
    - understaffing: Staffing below minimum on a day
    - conflict: Shift + absence conflict for an employee
    """
    from datetime import date as _date
    import calendar as _cal
    from collections import defaultdict

    today = _date.today()
    if year is None:
        year = today.year
    if month is None:
        month = today.month

    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="Ungültiger Monat: muss zwischen 1 und 12 liegen")

    db = get_db()
    warnings = []
    w_id = 0

    def make_id():
        nonlocal w_id
        w_id += 1
        return w_id

    # ── 1. Nächster Monat noch nicht geplant ─────────────────────
    # Check if current month → warn if < 7 days until month end and next month has no schedule
    last_day = _cal.monthrange(year, month)[1]
    month_end = _date(year, month, last_day)
    days_until_end = (month_end - today).days

    if days_until_end < 7:
        # Determine next month
        if month == 12:
            next_year, next_month = year + 1, 1
        else:
            next_year, next_month = year, month + 1

        next_prefix = f"{next_year:04d}-{next_month:02d}"
        next_month_mashi = [r for r in db._read("MASHI") if r.get("DATE", "").startswith(next_prefix)]
        next_month_spshi = [r for r in db._read("SPSHI") if r.get("DATE", "").startswith(next_prefix) and r.get("TYPE", 0) == 0]

        if len(next_month_mashi) + len(next_month_spshi) == 0:
            month_names_de = ["Januar", "Februar", "März", "April", "Mai", "Juni",
                               "Juli", "August", "September", "Oktober", "November", "Dezember"]
            next_month_name = month_names_de[next_month - 1]
            warnings.append({
                "id": make_id(),
                "type": "next_month_unplanned",
                "severity": "warning",
                "title": f"{next_month_name} {next_year} noch nicht geplant",
                "message": f"Nur noch {days_until_end} Tage bis Monatsende – der nächste Monat hat keinen Dienstplan.",
                "link": "/schedule",
                "link_label": "Zum Dienstplan",
                "date": today.isoformat(),
            })

    # ── 2. Überstunden > Schwellenwert ───────────────────────────
    OVERTIME_THRESHOLD = 20.0  # hours
    try:
        stats = db.get_statistics(year, month)
        for s in stats:
            ot = s.get("overtime_hours", 0)
            if ot > OVERTIME_THRESHOLD:
                warnings.append({
                    "id": make_id(),
                    "type": "overtime_exceeded",
                    "severity": "warning",
                    "title": f"Überstunden: {s['employee_name']}",
                    "message": f"{s['employee_name']} hat {ot:+.1f}h Überstunden in {year}/{month:02d}.",
                    "link": "/ueberstunden",
                    "link_label": "Zur Überstunden-Ansicht",
                    "employee_id": s["employee_id"],
                    "date": f"{year:04d}-{month:02d}-01",
                })
    except Exception:
        pass

    # ── 3. Besetzung unter Minimum ───────────────────────────────
    try:
        staffing_req = db.get_staffing_requirements()
        shift_reqs = staffing_req.get("shift_requirements", [])

        if shift_reqs:
            num_days = _cal.monthrange(year, month)[1]
            prefix = f"{year:04d}-{month:02d}"

            # Collect all schedule entries for the month once
            all_mashi = [r for r in db._read("MASHI") if r.get("DATE", "").startswith(prefix)]
            all_spshi = [r for r in db._read("SPSHI") if r.get("DATE", "").startswith(prefix) and r.get("TYPE", 0) == 0]

            for day_num in range(1, num_days + 1):
                from datetime import datetime as _datetime
                check_date = _datetime(year, month, day_num).date()
                check_str = check_date.isoformat()
                weekday = check_date.weekday()  # 0=Mon

                # Count by shift
                actual_by_shift: dict = defaultdict(int)
                for r in all_mashi:
                    if r.get("DATE", "") == check_str:
                        sid = r.get("SHIFTID")
                        if sid:
                            actual_by_shift[sid] += 1
                for r in all_spshi:
                    if r.get("DATE", "") == check_str:
                        sid = r.get("SHIFTID")
                        if sid:
                            actual_by_shift[sid] += 1

                for req in shift_reqs:
                    if req.get("weekday") != weekday:
                        continue
                    min_req = req.get("min", 0) or 0
                    if min_req == 0:
                        continue
                    shift_id = req.get("shift_id")
                    actual = actual_by_shift.get(shift_id, 0)
                    if actual < min_req:
                        shift_name = req.get("shift_name") or req.get("shift_short", "Schicht")
                        warnings.append({
                            "id": make_id(),
                            "type": "understaffing",
                            "severity": "error",
                            "title": f"Unterbesetzung: {shift_name} am {check_str}",
                            "message": f"Am {check_date.strftime('%d.%m.%Y')} fehlen {min_req - actual} Mitarbeiter für {shift_name} (Ist: {actual}, Soll: {min_req}).",
                            "link": "/schedule",
                            "link_label": "Zum Dienstplan",
                            "date": check_str,
                        })
    except Exception:
        pass

    # ── 4. Konflikte (Schicht + Abwesenheit) ─────────────────────
    try:
        conflicts = db.get_schedule_conflicts(year, month)
        for c in conflicts:
            warnings.append({
                "id": make_id(),
                "type": "conflict",
                "severity": "error",
                "title": f"Konflikt: {c['employee_name']}",
                "message": c.get("message", f"{c['employee_name']}: Schicht + Abwesenheit am {c['date']}"),
                "link": "/konflikte",
                "link_label": "Zu den Konflikten",
                "employee_id": c["employee_id"],
                "date": c["date"],
            })
    except Exception:
        pass

    return {
        "warnings": warnings,
        "count": len(warnings),
        "year": year,
        "month": month,
        "generated_at": today.isoformat(),
    }


# ── Fairness-Score ───────────────────────────────────────────────
@router.get(
    "/api/fairness",
    tags=["Statistics"],
    summary="Fairness analysis",
)
def get_fairness_score(
    year: int = Query(..., description="Year"),
    group_id: Optional[int] = Query(None, description="Filter by group"),
):
    """
    Berechnet den Fairness-Score: Wie gleichmäßig sind Wochenend-, Nacht-
    und Feiertagsschichten unter den Mitarbeitern verteilt?
    """
    import math

    db = get_db()
    employees = db.get_employees(include_hidden=False)
    if group_id:
        members = {m["employee_id"] for m in db.get_group_members(group_id)}
        employees = [e for e in employees if e["ID"] in members]

    shifts_map = {s["ID"]: s for s in db.get_shifts()}
    holidays_raw = db.get_holidays(year=year)
    holiday_dates = set()
    for h in holidays_raw:
        if isinstance(h, dict):
            d = h.get("DATE") or h.get("date")
            if d:
                holiday_dates.add(str(d)[:10])

    # Identify "night" shifts: start hour >= 20 or end hour <= 6
    def is_night(shift):
        t = shift.get("STARTEND0", "")
        if not t or "-" not in t:
            return False
        start = t.split("-")[0].strip()
        try:
            h = int(start.split(":")[0])
            return h >= 20 or h < 6
        except Exception:
            return False

    night_shift_ids = {sid for sid, s in shifts_map.items() if is_night(s)}

    # Collect all schedule entries for the year (month by month)
    all_entries: list[dict] = []
    for month in range(1, 13):
        entries = db.get_schedule(year=year, month=month, group_id=group_id)
        all_entries.extend(entries)

    # Count per employee
    stats: dict[int, dict] = {}
    for emp in employees:
        eid = emp["ID"]
        stats[eid] = {
            "employee_id": eid,
            "name": f"{emp.get('FIRSTNAME','')} {emp.get('NAME','')}".strip(),
            "shortname": emp.get("SHORTNAME", ""),
            "total": 0,
            "weekend": 0,
            "night": 0,
            "holiday": 0,
        }

    for entry in all_entries:
        eid = entry.get("employee_id")
        if eid not in stats:
            continue
        if entry.get("kind") != "shift":
            continue
        date_str = str(entry.get("date", ""))[:10]
        try:
            d = date.fromisoformat(date_str)
        except Exception:
            continue
        weekday = d.weekday()  # 0=Mo … 6=So
        stats[eid]["total"] += 1
        if weekday >= 5:
            stats[eid]["weekend"] += 1
        shift_id = entry.get("shift_id")
        if shift_id in night_shift_ids:
            stats[eid]["night"] += 1
        if date_str in holiday_dates:
            stats[eid]["holiday"] += 1

    result = [v for v in stats.values() if v["total"] > 0]
    if not result:
        return {"year": year, "employees": [], "fairness": {}}

    # Compute fairness metrics (std-dev based, lower = more fair)
    def score(values):
        if len(values) < 2:
            return 100.0
        mean = sum(values) / len(values)
        if mean == 0:
            return 100.0
        variance = sum((x - mean) ** 2 for x in values) / len(values)
        cv = math.sqrt(variance) / mean  # coefficient of variation
        return round(max(0, 100 - cv * 100), 1)

    weekend_vals = [r["weekend"] for r in result]
    night_vals   = [r["night"] for r in result]
    holiday_vals = [r["holiday"] for r in result]
    total_vals   = [r["total"] for r in result]

    fairness = {
        "weekend_score": score(weekend_vals),
        "night_score":   score(night_vals),
        "holiday_score": score(holiday_vals),
        "total_score":   score(total_vals),
        "overall":       round((score(weekend_vals) + score(night_vals) + score(total_vals)) / 3, 1),
        "avg_weekend":   round(sum(weekend_vals) / len(weekend_vals), 1),
        "avg_night":     round(sum(night_vals)   / len(night_vals),   1) if sum(night_vals) else 0,
        "avg_holiday":   round(sum(holiday_vals) / len(holiday_vals), 1) if sum(holiday_vals) else 0,
        "avg_total":     round(sum(total_vals)   / len(total_vals),   1),
    }

    result.sort(key=lambda x: x["name"])
    return {"year": year, "employees": result, "fairness": fairness}


# ── Kapazitäts-Forecast ──────────────────────────────────────────
@router.get(
    "/api/capacity-forecast",
    tags=["Statistics"],
    summary="Capacity forecast",
)
def get_capacity_forecast(
    year: int = Query(..., description="Year (YYYY)"),
    month: int = Query(..., description="Month (1-12)"),
    group_id: Optional[int] = Query(None, description="Filter by group"),
):
    """Return per-day capacity forecast for a month.

    Each day:
    - scheduled_count: employees with a shift entry
    - absent_count: employees with an absence
    - net_count: scheduled_count (absences are typically already subtracted from schedule)
    - absent_employees: list of {id, name, absence_type}
    - required_min: minimum requirement from staffing requirements
    - coverage_status: ok | low | critical | unknown
    - conflict_flag: True if absent_count is unusually high relative to total employees
    """
    import calendar as _cal
    from collections import defaultdict

    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="Ungültiger Monat: muss zwischen 1 und 12 liegen")

    db = get_db()
    num_days = _cal.monthrange(year, month)[1]
    prefix = f"{year:04d}-{month:02d}"

    # Get all employees (optionally filtered by group)
    all_employees = db.get_employees()
    if group_id:
        members = db.get_group_members(group_id)
        member_ids = {m.get('id', m.get('ID')) for m in members}
        all_employees = [e for e in all_employees if (e.get('id') or e.get('ID')) in member_ids]
    total_emp = len(all_employees)

    def _eid(e):
        return e.get('id') or e.get('ID')

    emp_by_id = {_eid(e): e for e in all_employees}

    # Build employee name lookup
    emp_name_by_id = {
        _eid(e): f"{e.get('firstname', e.get('FIRSTNAME',''))} {e.get('lastname', e.get('NAME',''))}".strip()
        for e in all_employees
    }

    # Get leave types for labels
    leave_types_list = db.get_leave_types()
    leave_label_by_id = {lt.get('id', lt.get('ID')): lt.get('short', lt.get('SHORTNAME', lt.get('name', lt.get('NAME', '?')))) for lt in leave_types_list}

    # Get staffing requirements for minimum thresholds
    staffing_req = db.get_staffing_requirements()
    shift_reqs = staffing_req.get("shift_requirements", [])

    # Build per-weekday minimum: max(min) across all shifts for that weekday
    min_by_weekday: dict = {}
    for req in shift_reqs:
        wd = req.get("weekday", -1)
        m = req.get("min", 0) or 0
        if wd >= 0:
            min_by_weekday[wd] = max(min_by_weekday.get(wd, 0), m)

    # Read schedule entries for the month
    all_mashi = [r for r in db._read("MASHI") if r.get("DATE", "").startswith(prefix)]
    all_spshi = [r for r in db._read("SPSHI") if r.get("DATE", "").startswith(prefix) and r.get("TYPE", 0) == 0]

    # Read absences for the month
    all_absences = [r for r in db._read("ABSEN") if r.get("DATE", "").startswith(prefix)]

    # Per-day aggregation
    day_scheduled: dict = defaultdict(set)  # day -> set of emp_ids
    day_absent: dict = defaultdict(list)    # day -> [{id, name, type}]

    for r in all_mashi:
        d = r.get("DATE", "")
        if d.startswith(prefix):
            try:
                day = int(d[8:10])
                eid = r.get("EMPLOYEEID")
                if eid and eid in emp_by_id:
                    day_scheduled[day].add(eid)
            except (ValueError, IndexError):
                pass

    for r in all_spshi:
        d = r.get("DATE", "")
        if d.startswith(prefix):
            try:
                day = int(d[8:10])
                eid = r.get("EMPLOYEEID")
                if eid and eid in emp_by_id:
                    day_scheduled[day].add(eid)
            except (ValueError, IndexError):
                pass

    for r in all_absences:
        d = r.get("DATE", "")
        if d.startswith(prefix):
            try:
                day = int(d[8:10])
                eid = r.get("EMPLOYEEID")
                if eid and eid in emp_by_id:
                    lt_id = r.get("LEAVETYPID") or r.get("LEAVETYPEID", 0)
                    lt_label = leave_label_by_id.get(lt_id, "Abw")
                    day_absent[day].append({
                        "id": eid,
                        "name": emp_name_by_id.get(eid, f"MA {eid}"),
                        "absence_type": lt_label,
                    })
            except (ValueError, IndexError):
                pass

    result = []
    for day in range(1, num_days + 1):
        import datetime as _dt
        check_date = _dt.date(year, month, day)
        weekday = check_date.weekday()  # 0=Mon

        scheduled = len(day_scheduled.get(day, set()))
        absent_list = day_absent.get(day, [])
        absent_count = len(absent_list)

        required_min = min_by_weekday.get(weekday, 0)

        # Coverage status
        if required_min > 0:
            diff = scheduled - required_min
            if diff >= 0:
                status = "ok"
            elif diff == -1:
                status = "low"
            else:
                status = "critical"
        else:
            # No requirement set — judge by absolute count
            if scheduled >= 3:
                status = "ok"
            elif scheduled >= 1:
                status = "low"
            elif scheduled == 0:
                status = "unplanned"
            else:
                status = "unknown"

        # Vacation conflict flag: more than 30% of team absent
        conflict_flag = total_emp > 0 and absent_count >= max(2, total_emp * 0.3)

        result.append({
            "day": day,
            "date": check_date.isoformat(),
            "weekday": weekday,  # 0=Mon
            "scheduled_count": scheduled,
            "absent_count": absent_count,
            "absent_employees": absent_list,
            "required_min": required_min,
            "coverage_status": status,
            "conflict_flag": conflict_flag,
            "total_employees": total_emp,
        })

    # Summary stats
    critical_days = [r for r in result if r["coverage_status"] == "critical"]
    low_days = [r for r in result if r["coverage_status"] == "low"]
    conflict_days = [r for r in result if r["conflict_flag"]]
    unplanned_days = [r for r in result if r["coverage_status"] == "unplanned"]

    return {
        "year": year,
        "month": month,
        "total_employees": total_emp,
        "days": result,
        "summary": {
            "critical_count": len(critical_days),
            "low_count": len(low_days),
            "conflict_count": len(conflict_days),
            "unplanned_count": len(unplanned_days),
            "ok_count": len([r for r in result if r["coverage_status"] == "ok"]),
        },
    }


# ── Jahres-Kapazitätsübersicht ───────────────────────────────────────────────
@router.get(
    "/api/capacity-year",
    tags=["Statistics"],
    summary="Yearly capacity overview",
)
def get_capacity_year(
    year: int = Query(..., description="Year (YYYY)"),
    group_id: Optional[int] = Query(None, description="Filter by group"),
    _cur_user: dict = Depends(require_auth),
):
    """Return per-month capacity summary for a full year (for heatmap).

    Each month:
    - avg_staffing: average daily staffing (planned days only)
    - ok_days, low_days, critical_days, unplanned_days
    - coverage_pct: avg_staffing / total_employees * 100
    - worst_status: overall month status
    """
    import calendar as _cal
    from collections import defaultdict

    db = get_db()
    all_employees = db.get_employees()
    if group_id:
        members = db.get_group_members(group_id)
        member_ids = {m.get('id', m.get('ID')) for m in members}
        all_employees = [e for e in all_employees if (e.get('id') or e.get('ID')) in member_ids]
    total_emp = len(all_employees)

    def _eid(e):
        return e.get('id') or e.get('ID')
    emp_by_id = {_eid(e): e for e in all_employees}

    staffing_req = db.get_staffing_requirements()
    shift_reqs = staffing_req.get("shift_requirements", [])
    min_by_weekday: dict = {}
    for req in shift_reqs:
        wd = req.get("weekday", -1)
        m = req.get("min", 0) or 0
        if wd >= 0:
            min_by_weekday[wd] = max(min_by_weekday.get(wd, 0), m)

    months_result = []
    for month in range(1, 13):
        import datetime as _dt
        num_days = _cal.monthrange(year, month)[1]
        prefix = f"{year:04d}-{month:02d}"

        all_mashi = [r for r in db._read("MASHI") if r.get("DATE", "").startswith(prefix)]
        all_spshi = [r for r in db._read("SPSHI") if r.get("DATE", "").startswith(prefix) and r.get("TYPE", 0) == 0]
        all_absences = [r for r in db._read("ABSEN") if r.get("DATE", "").startswith(prefix)]

        day_scheduled: dict = defaultdict(set)
        day_absent: dict = defaultdict(int)

        for r in all_mashi:
            d = r.get("DATE", "")
            if d.startswith(prefix):
                try:
                    day = int(d[8:10])
                    eid = r.get("EMPLOYEEID")
                    if eid and eid in emp_by_id:
                        day_scheduled[day].add(eid)
                except (ValueError, IndexError):
                    pass

        for r in all_spshi:
            d = r.get("DATE", "")
            if d.startswith(prefix):
                try:
                    day = int(d[8:10])
                    eid = r.get("EMPLOYEEID")
                    if eid and eid in emp_by_id:
                        day_scheduled[day].add(eid)
                except (ValueError, IndexError):
                    pass

        for r in all_absences:
            d = r.get("DATE", "")
            if d.startswith(prefix):
                try:
                    day = int(d[8:10])
                    day_absent[day] += 1
                except (ValueError, IndexError):
                    pass

        ok_days = low_days = critical_days = unplanned_days = 0
        staffing_sum = 0
        planned_days = 0

        for day in range(1, num_days + 1):
            check_date = _dt.date(year, month, day)
            weekday = check_date.weekday()
            scheduled = len(day_scheduled.get(day, set()))
            required_min = min_by_weekday.get(weekday, 0)

            if required_min > 0:
                diff = scheduled - required_min
                if diff >= 0:
                    ok_days += 1
                elif diff == -1:
                    low_days += 1
                else:
                    critical_days += 1
                staffing_sum += scheduled
                planned_days += 1
            else:
                if scheduled >= 3:
                    ok_days += 1
                    staffing_sum += scheduled
                    planned_days += 1
                elif scheduled >= 1:
                    low_days += 1
                    staffing_sum += scheduled
                    planned_days += 1
                else:
                    unplanned_days += 1

        avg_staffing = round(staffing_sum / max(1, planned_days), 1) if planned_days > 0 else 0
        coverage_pct = round(avg_staffing / max(1, total_emp) * 100, 1) if total_emp > 0 else 0

        if critical_days > 0:
            worst_status = "critical"
        elif low_days > 2:
            worst_status = "low"
        elif unplanned_days > num_days // 2:
            worst_status = "unplanned"
        else:
            worst_status = "ok"

        months_result.append({
            "month": month,
            "num_days": num_days,
            "avg_staffing": avg_staffing,
            "coverage_pct": coverage_pct,
            "ok_days": ok_days,
            "low_days": low_days,
            "critical_days": critical_days,
            "unplanned_days": unplanned_days,
            "planned_days": planned_days,
            "worst_status": worst_status,
            "total_employees": total_emp,
        })

    return {
        "year": year,
        "total_employees": total_emp,
        "months": months_result,
    }


# ── Qualitätsbericht ─────────────────────────────────────────────────────────

@router.get(
    "/api/quality-report",
    tags=["Statistics"],
    summary="Schedule quality report",
)
def get_quality_report(
    year: int = Query(...),
    month: int = Query(...),
):
    """Monatlicher Qualitätsbericht: Besetzung, Stunden-Compliance, Konflikte, Score."""
    import calendar as _cal
    from collections import defaultdict

    if not (1 <= month <= 12):
        raise HTTPException(status_code=400, detail="Ungültiger Monat: muss zwischen 1 und 12 liegen")

    db = get_db()
    num_days = _cal.monthrange(year, month)[1]
    prefix = f"{year:04d}-{month:02d}"
    month_name = _cal.month_name[month]

    # ── Mitarbeiter laden ───────────────────────────────────────────────────
    employees = {e["ID"]: e for e in db._read("EMPL") if not e.get("HIDE", 0)}
    active_emp_ids = set(employees.keys())

    # ── Schicht-Definitionen (Stunden) ──────────────────────────────────────
    shifts_by_id: dict = {}
    for s in db._read("SHIFT"):
        sid = s.get("ID")
        if sid:
            # Stunden = Dauer in h; DURATION in Minuten oder schon Stunden?
            dur_min = s.get("DURATION", 0)  # meist Minuten
            shifts_by_id[sid] = s.get("HOURS", dur_min / 60.0 if dur_min > 60 else dur_min)

    # ── Geplante Schichten (MASHI) ───────────────────────────────────────────
    day_emp_sets: dict = defaultdict(set)
    emp_actual_hours: dict = defaultdict(float)
    emp_shifts_count: dict = defaultdict(int)
    for r in db._read("MASHI"):
        d = r.get("DATE", "")
        if d.startswith(prefix):
            eid = r.get("EMPLOYEEID")
            if eid and eid in active_emp_ids:
                try:
                    day_num = int(d[8:10])
                    day_emp_sets[day_num].add(eid)
                    emp_shifts_count[eid] += 1
                    sid = r.get("SHIFTID")
                    hrs = shifts_by_id.get(sid, 8.0) if sid else 8.0
                    emp_actual_hours[eid] += hrs
                except (ValueError, IndexError):
                    pass

    # ── Abwesenheiten (ABSEN) ────────────────────────────────────────────────
    emp_absence_days: dict = defaultdict(int)
    emp_vacation_days: dict = defaultdict(int)
    emp_sick_days: dict = defaultdict(int)
    try:
        leave_types = {lt["ID"]: lt for lt in db._read("LEAVETYP")}
    except Exception:
        leave_types = {}
    for r in db._read("ABSEN"):
        d = r.get("DATE", "")
        if d.startswith(prefix):
            eid = r.get("EMPLOYEEID")
            if eid and eid in active_emp_ids:
                emp_absence_days[eid] += 1
                lt_id = r.get("LEAVETYPEID")
                lt = leave_types.get(lt_id, {})
                name = (lt.get("NAME") or lt.get("SHORTNAME") or "").lower()
                if "urlaub" in name or "vacation" in name or "holiday" in name:
                    emp_vacation_days[eid] += 1
                elif "krank" in name or "sick" in name:
                    emp_sick_days[eid] += 1

    # ── Tages-Besetzungs-Check ───────────────────────────────────────────────
    required_min = max(2, len(active_emp_ids) // 8)  # dynamisch
    coverage_days = []
    critical_days = []
    low_days = []
    ok_days = []
    unplanned_days = []

    for day in range(1, num_days + 1):
        date_obj = f"{year:04d}-{month:02d}-{day:02d}"
        wd = _cal.weekday(year, month, day)  # 0=Mon
        is_weekend = wd >= 5
        scheduled = len(day_emp_sets.get(day, set()))

        if scheduled == 0 and not is_weekend:
            status = "unplanned"
            unplanned_days.append(day)
        elif scheduled >= required_min:
            status = "ok"
            ok_days.append(day)
        elif scheduled == required_min - 1:
            status = "low"
            low_days.append(day)
        else:
            status = "critical"
            critical_days.append(day)

        coverage_days.append({
            "day": day,
            "date": date_obj,
            "weekday": wd,
            "is_weekend": is_weekend,
            "scheduled": scheduled,
            "required": required_min,
            "status": status,
        })

    # ── Stunden-Compliance (via get_statistics für korrekte Stunden-Berechnung) ──
    hours_issues = []
    hours_ok = []
    total_target = 0.0
    total_actual = 0.0

    try:
        stats_list = db.get_statistics(year, month)
    except Exception:
        stats_list = []

    for stat in stats_list:
        target = stat.get("target_hours") or 0.0
        actual = stat.get("actual_hours") or 0.0
        if target <= 0:
            continue
        total_target += target
        total_actual += actual
        deviation_pct = ((actual - target) / target * 100) if target else 0
        absence_days = stat.get("absence_days", 0) or 0
        issue_type = None
        if deviation_pct > 15:
            issue_type = "over"
        elif actual < target * 0.5 and absence_days < 10 and stat.get("shifts_count", 0) > 0:
            # stark unterstunden ohne Abwesenheiten = ungewöhnlich
            issue_type = "under"
        if issue_type:
            hours_issues.append({
                "employee_id": stat.get("employee_id"),
                "name": stat.get("employee_name", ""),
                "short": stat.get("employee_short", ""),
                "target_hours": round(target, 1),
                "actual_hours": round(actual, 1),
                "deviation_pct": round(deviation_pct, 1),
                "issue_type": issue_type,
                "shifts_count": stat.get("shifts_count", 0),
            })
        else:
            hours_ok.append(stat.get("employee_id"))

    # ── Score berechnen ──────────────────────────────────────────────────────
    # Formel: Gewichte Coverage 50%, Hours 30%, Konflikte 20%
    work_days_count = sum(1 for d in coverage_days if not d["is_weekend"])
    covered_work_days = sum(1 for d in coverage_days
                            if not d["is_weekend"] and d["status"] in ("ok", "low"))
    coverage_score = (covered_work_days / work_days_count * 100) if work_days_count else 100
    hours_score = max(0, 100 - len(hours_issues) * 5)
    # Conflicts: critical days & unplanned days penalty
    conflict_penalty = len(critical_days) * 5 + len(unplanned_days) * 3
    conflict_score = max(0, 100 - conflict_penalty)

    overall_score = round(coverage_score * 0.5 + hours_score * 0.3 + conflict_score * 0.2)
    if overall_score >= 90:
        grade = "A"
        grade_label = "Ausgezeichnet"
        grade_color = "green"
    elif overall_score >= 75:
        grade = "B"
        grade_label = "Gut"
        grade_color = "blue"
    elif overall_score >= 60:
        grade = "C"
        grade_label = "Verbesserungsbedarf"
        grade_color = "yellow"
    else:
        grade = "D"
        grade_label = "Kritisch"
        grade_color = "red"

    # ── Issues-Liste zusammenstellen ─────────────────────────────────────────
    findings = []
    if critical_days:
        findings.append({
            "severity": "critical",
            "category": "Besetzung",
            "message": f"{len(critical_days)} Tag(e) kritisch unterbesetzt",
            "days": critical_days,
        })
    if unplanned_days:
        findings.append({
            "severity": "warning",
            "category": "Planung",
            "message": f"{len(unplanned_days)} Werktag(e) ohne Planung",
            "days": unplanned_days,
        })
    if low_days:
        findings.append({
            "severity": "info",
            "category": "Besetzung",
            "message": f"{len(low_days)} Tag(e) knapp besetzt",
            "days": low_days,
        })
    over_emp = [h for h in hours_issues if h["issue_type"] == "over"]
    under_emp = [h for h in hours_issues if h["issue_type"] == "under"]
    if over_emp:
        findings.append({
            "severity": "warning",
            "category": "Überstunden",
            "message": f"{len(over_emp)} Mitarbeiter mit >15% Überstunden",
            "employees": [h["short"] for h in over_emp],
        })
    if under_emp:
        findings.append({
            "severity": "warning",
            "category": "Unterstunden",
            "message": f"{len(under_emp)} Mitarbeiter stark unterstundet",
            "employees": [h["short"] for h in under_emp],
        })
    if not findings:
        findings.append({
            "severity": "ok",
            "category": "Allgemein",
            "message": "Keine Auffälligkeiten — Monat kann abgeschlossen werden.",
        })

    return {
        "year": year,
        "month": month,
        "month_name": month_name,
        "overall_score": overall_score,
        "grade": grade,
        "grade_label": grade_label,
        "grade_color": grade_color,
        "active_employees": len(active_emp_ids),
        "work_days": work_days_count,
        "total_days": num_days,
        "required_min_per_day": required_min,
        "coverage": {
            "ok_days": len(ok_days),
            "low_days": len(low_days),
            "critical_days": len(critical_days),
            "unplanned_days": len(unplanned_days),
            "score": round(coverage_score, 1),
        },
        "hours": {
            "total_target": round(total_target, 1),
            "total_actual": round(total_actual, 1),
            "employees_ok": len(hours_ok),
            "employees_issues": len(hours_issues),
            "issues": hours_issues,
            "score": round(hours_score, 1),
        },
        "conflicts": {
            "score": round(conflict_score, 1),
            "critical_days": critical_days,
            "unplanned_days": unplanned_days,
        },
        "findings": findings,
        "coverage_days": coverage_days,
    }


# ── Verfügbarkeits-Matrix ────────────────────────────────────────────────────

@router.get(
    "/api/availability-matrix",
    tags=["Statistics"],
    summary="Employee availability matrix",
)
def get_availability_matrix(
    group_id: Optional[int] = Query(None),
    year: int = Query(None),
    months: int = Query(12, ge=1, le=24),
):
    """
    Analysiert Schicht-Muster aus dem Dienstplan (MASHI + SPSHI + ABSEN).
    Gibt pro Mitarbeiter zurück:
      - Schicht-Häufigkeit pro Wochentag (7 Tage × n Schichtarten)
      - Schicht-Mix (wie oft welche Schicht)
      - Muster-Label (z.B. "3-Schicht-Rotation", "Tagschicht Mo-Fr", "Frei")
    """
    import datetime
    from collections import defaultdict

    db = get_db()
    employees = db.get_employees(include_hidden=False)
    shifts_map = {s['ID']: s for s in db.get_shifts(include_hidden=True)}
    groups_map = {g['ID']: g['NAME'] for g in db.get_groups()}

    if group_id:
        members = set(db.get_group_members(group_id))
        employees = [e for e in employees if e['ID'] in members]

    if year is None:
        year = datetime.date.today().year

    # Collect all schedule entries for the requested range
    # month range: last `months` months up to end of `year`
    datetime.date.today()
    end_date = datetime.date(year, 12, 31)
    start_date = (end_date - datetime.timedelta(days=months * 30)).replace(day=1)

    # Build per-employee, per-weekday, per-shift counts
    # weekday: 0=Mo .. 6=So
    emp_wd_shift: dict[int, dict[int, dict]] = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
    emp_shift_total: dict[int, dict] = defaultdict(lambda: defaultdict(int))
    emp_day_total: dict[int, dict] = defaultdict(lambda: defaultdict(int))

    # Scan months
    cur = start_date.replace(day=1)
    while cur <= end_date:
        entries = db.get_schedule(cur.year, cur.month)
        for e in entries:
            d = datetime.date.fromisoformat(e['date'])
            if d < start_date or d > end_date:
                continue
            eid = e['employee_id']
            wd = d.weekday()  # 0=Mo
            sid = e.get('shift_id')
            if e.get('kind') in ('shift', 'special_shift') and sid:
                shift = shifts_map.get(sid)
                short = shift.get('SHORTNAME', '?') if shift else '?'
                name = shift.get('NAME', '?') if shift else '?'
                shift.get('COLORBK_HEX', '#888') if shift else '#888'
                emp_wd_shift[eid][wd][sid] += 1
                emp_shift_total[eid][sid] += 1
                emp_day_total[eid][wd] += 1
            elif e.get('kind') == 'absence':
                emp_wd_shift[eid][wd]['absence'] += 1
                emp_shift_total[eid]['absence'] += 1
                emp_day_total[eid][wd] += 1

        # advance month
        if cur.month == 12:
            cur = cur.replace(year=cur.year + 1, month=1)
        else:
            cur = cur.replace(month=cur.month + 1)

    # Build result per employee
    result_employees = []
    for emp in employees:
        eid = emp['ID']
        name = f"{emp.get('FIRSTNAME', '')} {emp.get('NAME', '')}".strip()
        short = emp.get('SHORTNAME', '')
        workdays = emp.get('WORKDAYS_LIST', [True] * 5 + [False, False])

        # Per-weekday breakdown
        weekday_data = []
        for wd in range(7):
            shift_counts = dict(emp_wd_shift[eid].get(wd, {}))
            total_for_day = emp_day_total[eid].get(wd, 0)
            # Build list of shifts sorted by count desc
            shifts_list = []
            for sid, cnt in sorted(shift_counts.items(), key=lambda x: -x[1]):
                if sid == 'absence':
                    shifts_list.append({
                        'shift_id': None,
                        'short': 'Ab',
                        'name': 'Abwesenheit',
                        'color': '#94a3b8',
                        'count': cnt,
                        'pct': round(cnt / total_for_day * 100) if total_for_day else 0,
                    })
                else:
                    shift = shifts_map.get(sid)
                    if shift:
                        shifts_list.append({
                            'shift_id': sid,
                            'short': shift.get('SHORTNAME', '?'),
                            'name': shift.get('NAME', '?'),
                            'color': shift.get('COLORBK_HEX', '#888'),
                            'count': cnt,
                            'pct': round(cnt / total_for_day * 100) if total_for_day else 0,
                        })
            weekday_data.append({
                'weekday': wd,
                'total': total_for_day,
                'configured': workdays[wd] if wd < len(workdays) else False,
                'shifts': shifts_list,
                # dominant shift
                'dominant_shift': shifts_list[0] if shifts_list else None,
            })

        # Overall shift mix
        total_shifts = sum(v for k, v in emp_shift_total[eid].items() if k != 'absence')
        shift_mix = []
        for sid, cnt in sorted(emp_shift_total[eid].items(), key=lambda x: -x[1]):
            if sid == 'absence':
                continue
            shift = shifts_map.get(sid)
            if shift:
                shift_mix.append({
                    'shift_id': sid,
                    'short': shift.get('SHORTNAME', '?'),
                    'name': shift.get('NAME', '?'),
                    'color': shift.get('COLORBK_HEX', '#888'),
                    'count': cnt,
                    'pct': round(cnt / total_shifts * 100) if total_shifts else 0,
                })

        # Pattern label
        active_wd = sum(1 for w in weekday_data if w['total'] > 0)
        if total_shifts == 0:
            pattern = 'Keine Daten'
            pattern_icon = '⬜'
        elif len(set(s['shift_id'] for s in shift_mix)) >= 3:
            pattern = '3-Schicht-Rotation'
            pattern_icon = '🔄'
        elif len(set(s['shift_id'] for s in shift_mix)) == 2:
            pattern = '2-Schicht-Wechsel'
            pattern_icon = '↔️'
        elif active_wd >= 5:
            pattern = 'Tagschicht Mo–Fr'
            pattern_icon = '☀️'
        elif active_wd >= 3:
            pattern = 'Teilzeit'
            pattern_icon = '📅'
        else:
            pattern = 'Wenige Einsätze'
            pattern_icon = '📉'

        # Group
        emp_groups = []
        for gid, gname in groups_map.items():
            members = set(db.get_group_members(gid))
            if eid in members:
                emp_groups.append(gname)

        result_employees.append({
            'id': eid,
            'name': name,
            'short': short,
            'groups': emp_groups,
            'pattern': pattern,
            'pattern_icon': pattern_icon,
            'total_shifts': total_shifts,
            'shift_mix': shift_mix,
            'weekdays': weekday_data,
            # configured workdays from employee record
            'workdays_config': workdays,
        })

    # Coverage per weekday (how many employees available)
    weekday_coverage = []
    for wd in range(7):
        configured = sum(1 for emp in employees
                         if wd < len(emp.get('WORKDAYS_LIST', [])) and emp['WORKDAYS_LIST'][wd])
        actual = sum(1 for e in result_employees if e['weekdays'][wd]['total'] > 0)
        weekday_coverage.append({
            'weekday': wd,
            'configured': configured,
            'actual_data': actual,
        })

    return {
        'year': year,
        'months': months,
        'start_date': start_date.isoformat(),
        'end_date': end_date.isoformat(),
        'employees': result_employees,
        'weekday_coverage': weekday_coverage,
        'shifts': [
            {'id': s['ID'], 'name': s['NAME'], 'short': s['SHORTNAME'], 'color': s.get('COLORBK_HEX', '#888')}
            for s in db.get_shifts(include_hidden=False)
        ],
    }


# ── Schichtplan-Simulation ────────────────────────────────────────────────────

class SimulationAbsence(BaseModel):
    emp_id: int = Field(..., gt=0)
    dates: list  # list of 'YYYY-MM-DD' strings, or ['all'] for whole month

class SimulationRequest(BaseModel):
    year: int = Field(..., ge=2000, le=2100)
    month: int = Field(..., ge=1, le=12)
    absences: list  # list of SimulationAbsence dicts
    scenario_name: Optional[str] = Field("Simulation", max_length=100)

@router.post(
    "/api/simulation",
    tags=["Statistics"],
    summary="Schedule simulation",
)
def run_simulation(body: SimulationRequest, _cur_user: dict = Depends(require_planer)):
    """
    Schichtplan-Simulation: Was passiert wenn Mitarbeiter ausfallen?
    Vergleicht Ist-Besetzung mit simulierter Besetzung nach Ausfall.
    """
    import calendar as _cal
    db = get_db()
    year, month = body.year, body.month
    entries = db.get_schedule(year=year, month=month)
    employees = db.get_employees(include_hidden=False)
    shifts = db.get_shifts(include_hidden=True)
    emp_map = {e['ID']: e for e in employees}
    {s['ID']: s for s in shifts}

    # Days in month
    days_in_month = _cal.monthrange(year, month)[1]
    date_strs = [f"{year}-{month:02d}-{d:02d}" for d in range(1, days_in_month + 1)]

    # Build absent set: {emp_id: set(dates)}
    absent_map: dict = {}
    for ab in body.absences:
        emp_id = ab['emp_id'] if isinstance(ab, dict) else ab.emp_id
        dates_raw = ab['dates'] if isinstance(ab, dict) else ab.dates
        if dates_raw == ['all'] or dates_raw == 'all':
            dates_set = set(date_strs)
        else:
            dates_set = set(dates_raw)
        absent_map[emp_id] = dates_set

    # Per-day analysis
    day_stats = []
    total_lost_shifts = 0
    critical_days = 0
    affected_employees = set(absent_map.keys())

    for date_str in date_strs:
        day_entries = [e for e in entries if e['date'] == date_str and e['kind'] == 'shift']
        baseline_count = len(day_entries)

        # Remove entries for absent employees on this date
        simulated = [
            e for e in day_entries
            if not (e['employee_id'] in absent_map and date_str in absent_map[e['employee_id']])
        ]
        sim_count = len(simulated)
        lost = baseline_count - sim_count
        total_lost_shifts += lost

        # Who is missing on this day
        missing_emps = []
        for e in day_entries:
            eid = e['employee_id']
            if eid in absent_map and date_str in absent_map[eid]:
                emp = emp_map.get(eid, {})
                missing_emps.append({
                    'emp_id': eid,
                    'name': f"{emp.get('FIRSTNAME','')} {emp.get('NAME','')}".strip(),
                    'shortname': emp.get('SHORTNAME', str(eid)),
                    'shift_id': e.get('shift_id'),
                    'shift_name': e.get('display_name', ''),
                })

        # Potential cover candidates: employees with shifts that day NOT absent
        cover_candidates = []
        for e in day_entries:
            eid = e['employee_id']
            if eid not in absent_map or date_str not in absent_map.get(eid, set()):
                emp = emp_map.get(eid, {})
                cover_candidates.append({
                    'emp_id': eid,
                    'name': f"{emp.get('FIRSTNAME','')} {emp.get('NAME','')}".strip(),
                    'shortname': emp.get('SHORTNAME', str(eid)),
                })

        # Status
        if sim_count == 0 and baseline_count > 0:
            status = 'critical'
            critical_days += 1
        elif lost > 0:
            status = 'degraded'
        else:
            status = 'ok'

        day_stats.append({
            'date': date_str,
            'day': int(date_str.split('-')[2]),
            'weekday': _cal.weekday(year, month, int(date_str.split('-')[2])),
            'baseline_count': baseline_count,
            'simulated_count': sim_count,
            'lost_shifts': lost,
            'status': status,
            'missing': missing_emps,
            'cover_candidates': cover_candidates[:5],  # top 5
        })

    # Per-employee impact summary
    employee_impacts = []
    for emp_id in absent_map:
        emp = emp_map.get(emp_id, {})
        emp_entries = [e for e in entries if e['employee_id'] == emp_id and e['kind'] == 'shift']
        absent_dates = absent_map[emp_id]
        affected = [e for e in emp_entries if e['date'] in absent_dates]
        employee_impacts.append({
            'emp_id': emp_id,
            'name': f"{emp.get('FIRSTNAME','')} {emp.get('NAME','')}".strip(),
            'shortname': emp.get('SHORTNAME', str(emp_id)),
            'total_shifts_in_month': len(emp_entries),
            'absent_shifts': len(affected),
            'absent_days': sorted(list(absent_dates & {e['date'] for e in emp_entries})),
        })

    return {
        'scenario_name': body.scenario_name,
        'year': year,
        'month': month,
        'days': day_stats,
        'summary': {
            'total_lost_shifts': total_lost_shifts,
            'critical_days': critical_days,
            'degraded_days': sum(1 for d in day_stats if d['status'] == 'degraded'),
            'ok_days': sum(1 for d in day_stats if d['status'] == 'ok'),
            'affected_employees': len(affected_employees),
        },
        'employee_impacts': employee_impacts,
    }
