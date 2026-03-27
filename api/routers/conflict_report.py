"""Conflict Report router — Q077.

Endpoints:
  GET  /api/v1/reports/conflicts         — detect shift conflicts in a date range
  GET  /api/v1/reports/conflicts/export  — export conflicts as CSV or XLSX

Conflict types detected:
  - overlap       : same employee, two shifts on the same day that overlap in time
  - double_booked : same employee assigned to two different shifts at the exact same time
  - understaffed  : a group has 0 employees scheduled on a day (but has members)
"""

from __future__ import annotations

import csv
import io
import logging
from datetime import date, timedelta

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import Response

from ..dependencies import get_db, limiter, require_planer

_logger = logging.getLogger("sp5api")

router = APIRouter()

# ── helpers ───────────────────────────────────────────────────────────────────


def _parse_time_str(t: str) -> int | None:
    """Parse 'HH:MM' → minutes from midnight.  Returns None on failure."""
    if not t:
        return None
    try:
        h, m = t.strip().split(":")
        return int(h) * 60 + int(m)
    except (ValueError, AttributeError):
        return None


def _parse_startend(startend: str) -> tuple[int, int] | None:
    """Parse 'HH:MM-HH:MM' → (start_min, end_min).

    Returns None if unparseable.  Handles overnight shifts.
    """
    if not startend or "-" not in startend:
        return None
    parts = startend.strip().split("-", 1)
    if len(parts) != 2:
        return None
    start = _parse_time_str(parts[0])
    end = _parse_time_str(parts[1])
    if start is None or end is None:
        return None
    if end <= start:
        end += 24 * 60  # overnight
    return (start, end)


def _ranges_overlap(a: tuple[int, int], b: tuple[int, int]) -> bool:
    return a[0] < b[1] and b[0] < a[1]


def _shift_time_range(shift: dict, weekday: int) -> tuple[int, int] | None:
    """Return (start_min, end_min) for a shift on a given weekday (0=Mon)."""
    key = f"STARTEND{weekday}"
    val = (shift.get(key) or "").strip()
    if val and "-" in val:
        r = _parse_startend(val)
        if r:
            return r
    # fallback to STARTEND0
    val0 = (shift.get("STARTEND0") or "").strip()
    return _parse_startend(val0)


def _date_range(from_date: date, to_date: date) -> list[date]:
    delta = (to_date - from_date).days + 1
    return [from_date + timedelta(days=i) for i in range(delta)]


# ── core detection logic ──────────────────────────────────────────────────────


def _detect_conflicts(
    db,
    from_date: date,
    to_date: date,
    group_id: int | None,
) -> list[dict]:
    """Run all conflict checks and return a list of conflict dicts."""
    conflicts: list[dict] = []

    # ── load reference data ──────────────────────────────────────────────────
    employees = {e["ID"]: e for e in db.get_employees(include_hidden=False)}
    shifts_map = {s["ID"]: s for s in db.get_shifts(include_hidden=True)}

    # group membership filter
    if group_id is not None:
        member_ids: set[int] | None = set(db.get_group_members(group_id))
    else:
        member_ids = None

    groups = {g["ID"]: g for g in db.get_groups()}

    # ── collect schedule entries (MASHI + SPSHI) in range ───────────────────
    from_str = from_date.isoformat()
    to_str = to_date.isoformat()

    # Build per-employee, per-date list of shifts: {(emp_id, date_str): [shift_id, ...]}
    emp_day_shifts: dict[tuple[int, str], list[int]] = {}

    # Read MASHI entries
    for r in db._read("MASHI"):
        d = r.get("DATE", "")
        if not d or not (from_str <= d[:10] <= to_str):
            continue
        eid = r.get("EMPLOYEEID")
        if eid is None or eid not in employees:
            continue
        if member_ids is not None and eid not in member_ids:
            continue
        sid = r.get("SHIFTID")
        if sid:
            key = (eid, d[:10])
            emp_day_shifts.setdefault(key, []).append(sid)

    # Read SPSHI entries (special shifts, TYPE==0 = shift, not absence)
    try:
        for r in db._read("SPSHI"):
            if r.get("TYPE", 0) != 0:
                continue
            d = r.get("DATE", "")
            if not d or not (from_str <= d[:10] <= to_str):
                continue
            eid = r.get("EMPLOYEEID")
            if eid is None or eid not in employees:
                continue
            if member_ids is not None and eid not in member_ids:
                continue
            sid = r.get("SHIFTID")
            if sid:
                key = (eid, d[:10])
                emp_day_shifts.setdefault(key, []).append(sid)
    except Exception:
        pass

    # ── check overlap + double_booked per employee/day ───────────────────────
    for (eid, date_str), shift_ids in emp_day_shifts.items():
        if len(shift_ids) < 2:
            continue

        emp = employees[eid]
        emp_name = f"{emp.get('FIRSTNAME', '')} {emp.get('NAME', '')}".strip()

        # determine weekday for time lookups
        try:
            d_obj = date.fromisoformat(date_str)
            wd = d_obj.weekday()  # 0=Mon
        except ValueError:
            wd = 0

        # build list of (shift_id, time_range_or_None)
        shift_ranges = []
        for sid in shift_ids:
            s = shifts_map.get(sid)
            tr = _shift_time_range(s, wd) if s else None
            shift_ranges.append((sid, tr))

        # pairwise checks
        for i in range(len(shift_ranges)):
            for j in range(i + 1, len(shift_ranges)):
                sid_a, range_a = shift_ranges[i]
                sid_b, range_b = shift_ranges[j]
                name_a = shifts_map.get(sid_a, {}).get("NAME", str(sid_a))
                name_b = shifts_map.get(sid_b, {}).get("NAME", str(sid_b))

                # Determine the group_id to report: use group_id param or find first group
                report_gid = group_id
                if report_gid is None:
                    for gid, _grp in groups.items():
                        mems = db.get_group_members(gid)
                        if eid in mems:
                            report_gid = gid
                            break

                if range_a is not None and range_b is not None:
                    if range_a[0] == range_b[0] and range_a[1] == range_b[1]:
                        # Identical time range = double_booked
                        conflicts.append({
                            "type": "double_booked",
                            "date": date_str,
                            "employee_id": eid,
                            "employee_name": emp_name,
                            "group_id": report_gid,
                            "description": (
                                f"{emp_name} is double-booked: "
                                f"'{name_a}' and '{name_b}' at the same time"
                            ),
                            "severity": "error",
                        })
                    elif _ranges_overlap(range_a, range_b):
                        conflicts.append({
                            "type": "overlap",
                            "date": date_str,
                            "employee_id": eid,
                            "employee_name": emp_name,
                            "group_id": report_gid,
                            "description": (
                                f"{emp_name}: shifts '{name_a}' and '{name_b}' "
                                "overlap on this day"
                            ),
                            "severity": "warning",
                        })
                else:
                    # No time data → can't determine overlap, flag as overlap warning
                    conflicts.append({
                        "type": "overlap",
                        "date": date_str,
                        "employee_id": eid,
                        "employee_name": emp_name,
                        "group_id": report_gid,
                        "description": (
                            f"{emp_name}: multiple shifts ('{name_a}' and '{name_b}') "
                            "assigned on this day"
                        ),
                        "severity": "warning",
                    })

    # ── understaffed days (group has members but 0 scheduled) ────────────────
    groups_to_check: list[tuple[int, set[int]]] = []
    if group_id is not None:
        mems = set(db.get_group_members(group_id))
        # Only filter to visible employees
        mems = mems & set(employees.keys())
        if mems:
            groups_to_check.append((group_id, mems))
    else:
        for gid, _grp in groups.items():
            mems = set(db.get_group_members(gid)) & set(employees.keys())
            if mems:
                groups_to_check.append((gid, mems))

    # build per-day scheduled set {date_str: set(emp_ids)} for the range
    day_scheduled: dict[str, set[int]] = {}
    for (eid, date_str), shift_ids in emp_day_shifts.items():
        if shift_ids:
            day_scheduled.setdefault(date_str, set()).add(eid)

    for d_obj in _date_range(from_date, to_date):
        date_str = d_obj.isoformat()
        scheduled_on_day = day_scheduled.get(date_str, set())
        for gid, mems in groups_to_check:
            group_name = groups.get(gid, {}).get("NAME", str(gid))
            # Only flag if no group member is scheduled at all
            if not (mems & scheduled_on_day):
                conflicts.append({
                    "type": "understaffed",
                    "date": date_str,
                    "employee_id": None,
                    "employee_name": None,
                    "group_id": gid,
                    "description": (
                        f"Group '{group_name}' has 0 employees scheduled on {date_str}"
                    ),
                    "severity": "warning",
                })

    # sort by date, then type
    conflicts.sort(key=lambda c: (c["date"], c["type"], c.get("employee_id") or 0))
    return conflicts


# ── summary helper ────────────────────────────────────────────────────────────

def _make_summary(conflicts: list[dict]) -> dict:
    return {
        "overlaps": sum(1 for c in conflicts if c["type"] == "overlap"),
        "double_booked": sum(1 for c in conflicts if c["type"] == "double_booked"),
        "understaffed": sum(1 for c in conflicts if c["type"] == "understaffed"),
    }


# ── GET /api/v1/reports/conflicts ─────────────────────────────────────────────


@router.get(
    "/api/reports/conflicts",
    tags=["Reports"],
    summary="Conflict report",
    description=(
        "Detect shift conflicts in a date range.\n\n"
        "Conflict types:\n"
        "- **overlap**: same employee has two shifts on the same day that overlap in time\n"
        "- **double_booked**: same employee assigned to two shifts at the exact same time\n"
        "- **understaffed**: a group has 0 employees scheduled on a day (but has members)\n\n"
        "**Required role:** Planer+  \n"
        "Also accessible via `/api/v1/reports/conflicts`."
    ),
    responses={
        200: {"description": "List of conflicts with summary"},
        400: {"description": "Invalid parameters"},
        401: {"description": "Not authenticated"},
    },
)
@limiter.limit("20/minute")
def get_conflict_report(
    request: Request,
    group_id: int | None = Query(None, description="Filter by group ID"),
    from_date: str = Query(..., alias="from", description="Start date YYYY-MM-DD"),
    to_date: str = Query(..., alias="to", description="End date YYYY-MM-DD"),
    _cur_user: dict = Depends(require_planer),
):
    try:
        f = date.fromisoformat(from_date)
        t = date.fromisoformat(to_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format — use YYYY-MM-DD")

    if t < f:
        raise HTTPException(status_code=400, detail="'to' must be >= 'from'")

    delta_days = (t - f).days
    if delta_days > 366:
        raise HTTPException(status_code=400, detail="Date range must not exceed 366 days")

    db = get_db()
    conflicts = _detect_conflicts(db, f, t, group_id)
    summary = _make_summary(conflicts)

    return {
        "conflicts": conflicts,
        "total": len(conflicts),
        "summary": summary,
    }


# ── GET /api/v1/reports/conflicts/export ──────────────────────────────────────


@router.get(
    "/api/reports/conflicts/export",
    tags=["Reports"],
    summary="Export conflict report",
    description=(
        "Export conflict report as CSV or XLSX.\n\n"
        "**Required role:** Planer+  \n"
        "Also accessible via `/api/v1/reports/conflicts/export`."
    ),
    responses={
        200: {"description": "File download"},
        400: {"description": "Invalid parameters"},
        401: {"description": "Not authenticated"},
    },
)
@limiter.limit("10/minute")
def export_conflict_report(
    request: Request,
    group_id: int | None = Query(None, description="Filter by group ID"),
    from_date: str = Query(..., alias="from", description="Start date YYYY-MM-DD"),
    to_date: str = Query(..., alias="to", description="End date YYYY-MM-DD"),
    format: str = Query("csv", description="Output format: csv or xlsx"),
    _cur_user: dict = Depends(require_planer),
):
    try:
        f = date.fromisoformat(from_date)
        t = date.fromisoformat(to_date)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format — use YYYY-MM-DD")

    if t < f:
        raise HTTPException(status_code=400, detail="'to' must be >= 'from'")
    if format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="format must be 'csv' or 'xlsx'")

    delta_days = (t - f).days
    if delta_days > 366:
        raise HTTPException(status_code=400, detail="Date range must not exceed 366 days")

    db = get_db()
    conflicts = _detect_conflicts(db, f, t, group_id)

    # Build flat rows for export
    rows = []
    for c in conflicts:
        rows.append({
            "Type": c["type"],
            "Severity": c["severity"],
            "Date": c["date"],
            "Employee ID": c.get("employee_id") or "",
            "Employee": c.get("employee_name") or "",
            "Group ID": c.get("group_id") or "",
            "Description": c["description"],
        })

    filename_base = f"conflicts_{from_date}_{to_date}"

    if format == "csv":
        buf = io.StringIO()
        if rows:
            writer = csv.DictWriter(buf, fieldnames=rows[0].keys(), lineterminator="\r\n")
            writer.writeheader()
            writer.writerows(rows)
        content = ("\ufeff" + buf.getvalue()).encode("utf-8")
        return Response(
            content=content,
            media_type="text/csv; charset=utf-8-sig",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    # xlsx
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conflicts"

    thin = Side(border_style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if rows:
        headers = list(rows[0].keys())
        col_widths = [16, 12, 14, 14, 28, 12, 60]
        for c_idx, (h, w) in enumerate(zip(headers, col_widths, strict=False), start=1):
            cell = ws.cell(1, c_idx, h)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill(fill_type="solid", fgColor="1E293B")
            cell.alignment = Alignment(horizontal="left")
            cell.border = border
            ws.column_dimensions[get_column_letter(c_idx)].width = w

        severity_colors = {"error": "FEE2E2", "warning": "FEF9C3"}
        for r_idx, row in enumerate(rows, start=2):
            fill_color = severity_colors.get(row["Severity"], "FFFFFF")
            for c_idx, val in enumerate(row.values(), start=1):
                cell = ws.cell(r_idx, c_idx, val)
                cell.font = Font(size=9)
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
                cell.border = border
                cell.alignment = Alignment(wrap_text=(c_idx == len(headers)))
        ws.freeze_panes = "A2"

    buf_bytes = io.BytesIO()
    wb.save(buf_bytes)
    return Response(
        content=buf_bytes.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'},
    )
